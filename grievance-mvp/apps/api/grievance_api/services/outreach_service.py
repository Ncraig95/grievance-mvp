from __future__ import annotations

import asyncio
import base64
import csv
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
import hashlib
import html
from io import BytesIO, StringIO
import json
import logging
from pathlib import Path
import re
import secrets
from typing import Any
from urllib.parse import quote, urlsplit
from zoneinfo import ZoneInfo

import aiosqlite
from jinja2 import ChainableUndefined, Environment, meta
from markupsafe import Markup
from openpyxl import load_workbook

from ..core.config import EmailConfig, OfficerAuthConfig, OutreachConfig
from ..db.db import Db, utcnow
from .graph_mail import GraphMailer


_BOOL_TRUE = {"1", "true", "yes", "y", "on"}
_URL_RE = re.compile(r"(https?://[^\s<]+)")
_HTML_HREF_RE = re.compile(r'href="([^"]+)"')
_NORMALIZE_KEY_RE = re.compile(r"[^a-z0-9]+")
_MULTI_VALUE_SPLIT_RE = re.compile(r"[\r\n,;|]+")
_OUTREACH_HTML_TEMPLATE_FILENAME = "outreach_base.html"
_STRONG_AUTOMATION_HINTS = (
    "barracuda",
    "mimecast",
    "proofpoint",
    "safelinks",
    "urlscan",
    "security",
    "crawler",
    "spider",
    "bot",
    "prefetch",
    "headless",
    "python-requests",
    "curl",
    "wget",
)
_PIXEL_GIF_BYTES = base64.b64decode("R0lGODlhAQABAPAAAAAAAAAAACH5BAEAAAAALAAAAAABAAEAAAICRAEAOw==")

_CONTACT_HEADER_ALIASES: dict[str, str] = {
    "email": "email",
    "email_address": "email",
    "mail": "email",
    "first_name": "first_name",
    "firstname": "first_name",
    "last_name": "last_name",
    "lastname": "last_name",
    "full_name": "full_name",
    "fullname": "full_name",
    "name": "full_name",
    "location": "work_location",
    "work_location": "work_location",
    "workgroup": "work_group",
    "work_group": "work_group",
    "group": "group_name",
    "group_name": "group_name",
    "subgroup": "subgroup_name",
    "sub_group": "subgroup_name",
    "subgroup_name": "subgroup_name",
    "department": "department",
    "bargaining_unit": "bargaining_unit",
    "bargainingunit": "bargaining_unit",
    "local_number": "local_number",
    "localnumber": "local_number",
    "steward_name": "steward_name",
    "steward": "steward_name",
    "rep_name": "rep_name",
    "assigned_rep": "rep_name",
    "officer_name": "rep_name",
    "notes": "notes",
    "active": "active",
}

_IMPORT_FIELD_KEYS: tuple[str, ...] = (
    "email",
    "first_name",
    "last_name",
    "full_name",
    "work_location",
    "work_group",
    "group_name",
    "subgroup_name",
    "department",
    "bargaining_unit",
    "local_number",
    "steward_name",
    "rep_name",
)

_GROUP_HEADER_PRIORITY: tuple[str, ...] = (
    "group_name",
    "group",
    "work_group",
    "workgroup",
    "processing_unit",
    "field_unit",
    "unit",
    "bargaining_unit",
    "bargainingunit",
    "member_department",
    "department",
    "field_department",
    "branch",
    "region",
    "division",
    "field_division",
    "field_employer",
    "work_location",
    "customworklocation1",
    "customworklocation2",
)

_SUBGROUP_HEADER_PRIORITY: tuple[str, ...] = (
    "subgroup_name",
    "subgroup",
    "sub_group",
    "division",
    "field_division",
    "member_department",
    "department",
    "field_department",
    "branch",
    "region",
    "work_location",
    "customworklocation1",
    "customworklocation2",
    "job_title_description",
)

_COMBINED_STATUS_HEADER_ALIASES: set[str] = {
    "status",
    "member_status",
    "membership_status",
    "member_status_combined",
    "status_bucket",
    "member_type_status",
}

_MEMBERSHIP_STATUS_HEADER_ALIASES: set[str] = {
    "membership_type",
    "member_type",
    "member_non_member",
    "member_status_type",
    "member_category",
}

_EMPLOYMENT_STATUS_HEADER_ALIASES: set[str] = {
    "employment_status",
    "employment",
    "active_status",
    "worker_status",
    "activity_status",
}

_STATUS_DETAIL_HEADER_ALIASES: set[str] = {
    "status_detail",
    "detail_status",
    "detail",
    "member_detail",
    "status_reason",
}

_ALLOWED_STATUS_BUCKETS: tuple[str, ...] = (
    "Member - Active - Active",
    "Member - Active - Pending",
    "Non Member - Active - Active",
    "Non Member - Active - Non fr Mem",
)

_PLACEHOLDER_CATALOG: tuple[str, ...] = (
    "first_name",
    "last_name",
    "full_name",
    "email",
    "location",
    "campaign_location",
    "work_location",
    "work_group",
    "group_name",
    "subgroup_name",
    "department",
    "bargaining_unit",
    "local_number",
    "steward_name",
    "rep_name",
    "membership_type",
    "employment_status",
    "status_detail",
    "status_bucket",
    "status_source_text",
    "visit_date",
    "visit_time",
    "subject",
    "sender_name",
    "reply_to",
    "unsubscribe_url",
)

_SEEDED_TEMPLATES: tuple[dict[str, str], ...] = (
    {
        "template_key": "outreach_notice",
        "name": "Initial Notice",
        "template_type": "notice",
        "subject_template": "{{ location }} Visit on {{ visit_date }}",
        "body_template": """Hi {{ first_name | default('everyone') }},

We wanted to let you know that we will be at {{ location }} on {{ visit_date }} from {{ visit_time }} for organizing outreach.

Please stop by and talk with us before work, on your lunch break, during your 15-minute break, or after work if you have time. This is a good chance to ask questions, share workplace concerns, learn more about what the union is working on, and let us know what issues matter most to you.

Whether you have a specific question, want an update, or just want to meet with us face to face, we would really like to hear from you. Your input helps us better represent everyone in your area.

We hope to see you there.

Thank you,
Nick""",
    },
    {
        "template_key": "outreach_reminder",
        "name": "Day-Of Reminder",
        "template_type": "reminder",
        "subject_template": "Reminder: {{ location }} Visit on {{ visit_date }}",
        "body_template": """Hi {{ first_name | default('everyone') }},

This is a reminder that we will be at {{ location }} today, {{ visit_date }}, from {{ visit_time }}.

If you have a few minutes, please come see us before work, on your lunch break, during your 15-minute break, or after work. We are there to answer questions, hear concerns, share updates, and make sure you have a chance to connect with the union directly.

Even a quick conversation can help. We want to hear what is going well, what needs attention, and how we can better support employees in your area.

We hope to see you today.

Thank you,
Nick""",
    },
)

_SEEDED_STOPS: tuple[dict[str, str], ...] = (
    {"location_name": "Ed Ball Building", "visit_date_local": "2026-04-14", "start_time_local": "08:00", "end_time_local": "14:00", "notice_subject": "Ed Ball Building Visit on April 14th", "reminder_subject": "Ed Ball Building Visit on April 14th"},
    {"location_name": "Ed Ball Building", "visit_date_local": "2026-05-05", "start_time_local": "08:00", "end_time_local": "14:00", "notice_subject": "Ed Ball Building Visit on May 5th", "reminder_subject": "Ed Ball Building Visit on May 5th"},
    {"location_name": "Police Memorial Building", "visit_date_local": "2026-04-15", "start_time_local": "10:00", "end_time_local": "13:00", "notice_subject": "Police Memorial Building Visit on April 15th", "reminder_subject": "Police Memorial Building Visit on April 15th"},
    {"location_name": "Police Memorial Building", "visit_date_local": "2026-05-06", "start_time_local": "10:00", "end_time_local": "13:00", "notice_subject": "Police Memorial Building Visit on May 6th", "reminder_subject": "Police Memorial Building Visit on May 6th"},
    {"location_name": "Main Library", "visit_date_local": "2026-04-16", "start_time_local": "10:00", "end_time_local": "14:00", "notice_subject": "Main Library Visit on April 16th", "reminder_subject": "Main Library Visit on April 16th"},
    {"location_name": "Main Library", "visit_date_local": "2026-05-07", "start_time_local": "10:00", "end_time_local": "14:00", "notice_subject": "Main Library Visit on May 7th", "reminder_subject": "Main Library Visit on May 7th"},
    {"location_name": "Fire and Rescue", "visit_date_local": "2026-04-21", "start_time_local": "10:00", "end_time_local": "13:00", "notice_subject": "Fire and Rescue Visit on April 21st", "reminder_subject": "Fire and Rescue Visit on April 21st"},
    {"location_name": "Fire and Rescue", "visit_date_local": "2026-05-13", "start_time_local": "10:00", "end_time_local": "13:00", "notice_subject": "Fire and Rescue Orientation Visit on May 13th", "reminder_subject": "Fire and Rescue Orientation Visit on May 13th"},
    {"location_name": "Art Museum, Behavioral and Human Services", "visit_date_local": "2026-04-22", "start_time_local": "15:00", "end_time_local": "16:30", "notice_subject": "Art Museum, Behavioral and Human Services Visit on April 22nd", "reminder_subject": "Art Museum, Behavioral and Human Services Visit on April 22nd"},
    {"location_name": "Art Museum, Behavioral and Human Services", "visit_date_local": "2026-05-12", "start_time_local": "15:00", "end_time_local": "16:30", "notice_subject": "Art Museum, Behavioral and Human Services Visit on May 12th", "reminder_subject": "Art Museum, Behavioral and Human Services Visit on May 12th"},
    {"location_name": "City Hall", "visit_date_local": "2026-04-23", "start_time_local": "10:00", "end_time_local": "15:00", "notice_subject": "City Hall Visit on April 23rd", "reminder_subject": "City Hall Visit on April 23rd"},
    {"location_name": "City Hall", "visit_date_local": "2026-05-14", "start_time_local": "10:00", "end_time_local": "15:00", "notice_subject": "City Hall Visit on May 14th", "reminder_subject": "City Hall Visit on May 14th"},
    {"location_name": "Pablo Branch", "visit_date_local": "2026-04-28", "start_time_local": "15:00", "end_time_local": "16:30", "notice_subject": "Pablo Branch Visit on April 28th", "reminder_subject": "Pablo Branch Visit on April 28th"},
    {"location_name": "Pablo Branch", "visit_date_local": "2026-05-19", "start_time_local": "15:00", "end_time_local": "16:30", "notice_subject": "Pablo Branch Visit on May 19th", "reminder_subject": "Pablo Branch Visit on May 19th"},
    {"location_name": "University Branch", "visit_date_local": "2026-04-29", "start_time_local": "15:00", "end_time_local": "16:30", "notice_subject": "University Branch Visit on April 29th", "reminder_subject": "University Branch Visit on April 29th"},
    {"location_name": "University Branch", "visit_date_local": "2026-05-20", "start_time_local": "15:00", "end_time_local": "16:30", "notice_subject": "University Branch Visit on May 20th", "reminder_subject": "University Branch Visit on May 20th"},
    {"location_name": "South Branch", "visit_date_local": "2026-04-30", "start_time_local": "15:00", "end_time_local": "16:30", "notice_subject": "South Branch Visit on April 30th", "reminder_subject": "South Branch Visit on April 30th"},
    {"location_name": "South Branch", "visit_date_local": "2026-05-21", "start_time_local": "15:00", "end_time_local": "16:30", "notice_subject": "South Branch Visit on May 21st", "reminder_subject": "South Branch Visit on May 21st"},
)

_DEFAULT_OUTREACH_HTML_TEMPLATE = """<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ subject }}</title>
  </head>
  <body style="margin:0;padding:0;background-color:#f5efe6;">
    <div style="display:none;max-height:0;overflow:hidden;opacity:0;color:transparent;">
      {{ subject }}
    </div>
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;background-color:#f5efe6;">
      <tr>
        <td align="center" style="padding:24px 12px;">
          <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;max-width:680px;background-color:#ffffff;border:1px solid #d6d3d1;">
            <tr>
              <td style="padding:28px 32px 24px;background-color:#0f3d33;color:#f8fafc;font-family:Segoe UI,Tahoma,Geneva,Verdana,sans-serif;">
                <div style="font-size:12px;line-height:1.4;letter-spacing:0.12em;text-transform:uppercase;opacity:0.88;">
                  {{ sender_name | default('CWA Local 3106') }}
                </div>
                <div style="margin-top:10px;font-size:28px;line-height:1.2;font-weight:700;">
                  {{ subject }}
                </div>
              </td>
            </tr>
            <tr>
              <td style="padding:32px 32px 18px;color:#1f2937;font-family:Segoe UI,Tahoma,Geneva,Verdana,sans-serif;font-size:16px;line-height:1.65;">
                {{ content_html }}
              </td>
            </tr>
            <tr>
              <td style="padding:0 32px 32px;font-family:Segoe UI,Tahoma,Geneva,Verdana,sans-serif;">
                <table role="presentation" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
                  <tr>
                    <td style="border-radius:999px;background-color:#0f3d33;">
                      <a href="{{ unsubscribe_url }}" style="display:inline-block;padding:12px 20px;color:#ffffff;font-size:14px;font-weight:700;text-decoration:none;">
                        Unsubscribe
                      </a>
                    </td>
                  </tr>
                </table>
                <p style="margin:18px 0 0;font-size:12px;line-height:1.6;color:#57534e;">
                  If the button does not open, use this link:
                  <br>
                  <a href="{{ unsubscribe_url }}" style="color:#0f766e;text-decoration:underline;word-break:break-all;">{{ unsubscribe_url }}</a>
                </p>
                {% if reply_to %}
                <p style="margin:14px 0 0;font-size:12px;line-height:1.6;color:#57534e;">
                  Reply to
                  <a href="mailto:{{ reply_to }}" style="color:#0f766e;text-decoration:underline;">{{ reply_to }}</a>
                </p>
                {% endif %}
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""


@dataclass(frozen=True)
class OutreachRenderedMessage:
    subject: str
    text_body: str
    html_body: str
    unknown_placeholders: list[str]


@dataclass(frozen=True)
class OutreachSendSummary:
    send_log_id: int
    recipient_email: str
    status: str
    graph_message_id: str | None = None
    error_text: str | None = None


@dataclass(frozen=True)
class OutreachImportSheet:
    name: str
    headers: list[str]
    rows: list[dict[str, str]]


def _normalize_key(value: object) -> str:
    return _NORMALIZE_KEY_RE.sub("_", str(value or "").strip().lower()).strip("_")


def _normalize_text(value: object) -> str:
    return str(value or "").strip()


def _normalize_email(value: object) -> str:
    return _normalize_text(value).lower()


def _normalize_text_list(value: object) -> list[str]:
    seen: set[str] = set()
    tokens: list[str] = []
    for raw_token in _MULTI_VALUE_SPLIT_RE.split(_normalize_text(value)):
        token = " ".join(_normalize_text(raw_token).split())
        if not token:
            continue
        normalized_key = token.lower()
        if normalized_key in seen:
            continue
        seen.add(normalized_key)
        tokens.append(token)
    return tokens


def _normalize_text_list_value(value: object) -> str | None:
    tokens = _normalize_text_list(value)
    if not tokens:
        return None
    return ", ".join(tokens)


def _full_name(first_name: str, last_name: str, fallback: str = "") -> str:
    joined = " ".join(part for part in (first_name.strip(), last_name.strip()) if part).strip()
    return joined or fallback.strip()


def _json_loads(raw_value: object, *, default: dict[str, Any] | None = None) -> dict[str, Any]:
    text = _normalize_text(raw_value)
    if not text:
        return dict(default or {})
    try:
        parsed = json.loads(text)
    except Exception:
        return dict(default or {})
    if isinstance(parsed, dict):
        return parsed
    return dict(default or {})


def _as_bool(value: object, *, default: bool = True) -> bool:
    text = _normalize_text(value).lower()
    if not text:
        return default
    return text in _BOOL_TRUE


def _token_hash(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def _hash_text(value: object) -> str | None:
    text = _normalize_text(value)
    if not text:
        return None
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _header_fingerprint(headers: list[str]) -> str:
    normalized = sorted(_normalize_key(header) for header in headers if _normalize_key(header))
    return hashlib.sha256(json.dumps(normalized, separators=(",", ":"), ensure_ascii=True).encode("utf-8")).hexdigest()


def _normalized_headers(headers: list[str]) -> list[str]:
    return [_normalize_key(header) for header in headers if _normalize_key(header)]


def _header_value_stats(rows: list[dict[str, str]], header: str) -> tuple[int, int]:
    nonblank_values = [_normalize_text(row.get(header)) for row in rows if _normalize_text(row.get(header))]
    return len(nonblank_values), len({value for value in nonblank_values})


def _sanitize_mapping_dict(raw_value: object) -> dict[str, str | None]:
    if not isinstance(raw_value, dict):
        return {}
    out: dict[str, str | None] = {}
    for key, value in raw_value.items():
        normalized_key = _normalize_key(key)
        if not normalized_key:
            continue
        out[normalized_key] = _normalize_text(value) or None
    return out


def _canonical_membership_type(value: object) -> str | None:
    normalized = _normalize_key(value)
    if normalized in {"member", "mem"}:
        return "Member"
    if normalized in {"non_member", "nonmember", "non_member_", "nonmember_", "non_member_member"}:
        return "Non Member"
    return None


def _canonical_employment_status(value: object) -> str | None:
    normalized = _normalize_key(value)
    if normalized == "active":
        return "Active"
    return None


def _canonical_status_detail(value: object) -> str | None:
    normalized = _normalize_key(value)
    if normalized == "active":
        return "Active"
    if normalized == "pending":
        return "Pending"
    if normalized in {"non_fr_mem", "nonfrmem", "non_fr_member"}:
        return "Non fr Mem"
    return None


def _status_bucket(
    *,
    membership_type: str | None,
    employment_status: str | None,
    status_detail: str | None,
) -> str | None:
    if not membership_type or not employment_status or not status_detail:
        return None
    bucket = f"{membership_type} - {employment_status} - {status_detail}"
    if bucket in _ALLOWED_STATUS_BUCKETS:
        return bucket
    return None


def _classify_combined_status(value: object) -> dict[str, str | None]:
    text = _normalize_text(value)
    if not text:
        return {
            "membership_type": None,
            "employment_status": None,
            "status_detail": None,
            "status_bucket": None,
            "status_source_text": None,
            "reason": "missing_status",
        }
    parts = [part.strip() for part in re.split(r"\s+-\s+", text) if part.strip()]
    if len(parts) != 3:
        return {
            "membership_type": None,
            "employment_status": None,
            "status_detail": None,
            "status_bucket": None,
            "status_source_text": text,
            "reason": "unsupported_status",
        }
    membership_type = _canonical_membership_type(parts[0])
    employment_status = _canonical_employment_status(parts[1])
    status_detail = _canonical_status_detail(parts[2])
    bucket = _status_bucket(
        membership_type=membership_type,
        employment_status=employment_status,
        status_detail=status_detail,
    )
    return {
        "membership_type": membership_type,
        "employment_status": employment_status,
        "status_detail": status_detail,
        "status_bucket": bucket,
        "status_source_text": text,
        "reason": None if bucket else "unsupported_status",
    }


def _classify_split_status(
    *,
    membership_type_value: object,
    employment_status_value: object,
    status_detail_value: object,
) -> dict[str, str | None]:
    source_text = " - ".join(
        value for value in (
            _normalize_text(membership_type_value),
            _normalize_text(employment_status_value),
            _normalize_text(status_detail_value),
        )
        if value
    )
    if not source_text:
        return {
            "membership_type": None,
            "employment_status": None,
            "status_detail": None,
            "status_bucket": None,
            "status_source_text": None,
            "reason": "missing_status",
        }
    if not (
        _normalize_text(membership_type_value)
        and _normalize_text(employment_status_value)
        and _normalize_text(status_detail_value)
    ):
        return {
            "membership_type": None,
            "employment_status": None,
            "status_detail": None,
            "status_bucket": None,
            "status_source_text": source_text,
            "reason": "incomplete_status",
        }
    membership_type = _canonical_membership_type(membership_type_value)
    employment_status = _canonical_employment_status(employment_status_value)
    status_detail = _canonical_status_detail(status_detail_value)
    bucket = _status_bucket(
        membership_type=membership_type,
        employment_status=employment_status,
        status_detail=status_detail,
    )
    return {
        "membership_type": membership_type,
        "employment_status": employment_status,
        "status_detail": status_detail,
        "status_bucket": bucket,
        "status_source_text": source_text,
        "reason": None if bucket else "unsupported_status",
    }


def _parse_local_date(value: str) -> date:
    return date.fromisoformat(_normalize_text(value))


def _parse_local_time(value: str) -> time:
    text = _normalize_text(value)
    if len(text) == 5:
        return time.fromisoformat(text)
    return time.fromisoformat(text[:8])


def _ordinal_day(day: int) -> str:
    if 10 <= day % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
    return f"{day}{suffix}"


def _friendly_visit_date(value: str) -> str:
    parsed = _parse_local_date(value)
    return parsed.strftime("%B ") + _ordinal_day(parsed.day)


def _friendly_visit_time(start_value: str, end_value: str) -> str:
    start_dt = datetime.combine(date(2000, 1, 1), _parse_local_time(start_value))
    end_dt = datetime.combine(date(2000, 1, 1), _parse_local_time(end_value))
    return f"{start_dt.strftime('%-I:%M %p')} to {end_dt.strftime('%-I:%M %p')}"


def _local_to_utc_iso(*, tz_name: str, local_date: str, local_time_value: time) -> str:
    zone = ZoneInfo(tz_name)
    local_dt = datetime.combine(_parse_local_date(local_date), local_time_value).replace(tzinfo=zone)
    return local_dt.astimezone(timezone.utc).isoformat()


def _utc_to_local_input(utc_value: str, tz_name: str) -> str:
    zone = ZoneInfo(tz_name)
    return datetime.fromisoformat(utc_value).astimezone(zone).strftime("%Y-%m-%dT%H:%M")


def _default_notice_send_at(*, visit_date_local: str, tz_name: str) -> str:
    return _local_to_utc_iso(
        tz_name=tz_name,
        local_date=(_parse_local_date(visit_date_local) - timedelta(days=4)).isoformat(),
        local_time_value=time(8, 30),
    )


def _default_reminder_send_at(*, visit_date_local: str, start_time_local: str, tz_name: str) -> str:
    start_local = _parse_local_time(start_time_local)
    reminder_time = time(11, 30) if start_local >= time(15, 0) else time(6, 30)
    return _local_to_utc_iso(
        tz_name=tz_name,
        local_date=visit_date_local,
        local_time_value=reminder_time,
    )


def _public_base_url(outreach_cfg: OutreachConfig, email_cfg: EmailConfig, officer_auth_cfg: OfficerAuthConfig) -> str:
    explicit = _normalize_text(outreach_cfg.public_base_url).rstrip("/")
    if explicit:
        return explicit
    for candidate in (email_cfg.approval_request_url_base, officer_auth_cfg.redirect_uri):
        parsed = urlsplit(_normalize_text(candidate))
        if parsed.scheme and parsed.netloc:
            return f"{parsed.scheme}://{parsed.netloc}"
    return ""


def _normalize_user_agent(value: object) -> str:
    return " ".join(_normalize_text(value).split())


def _normalize_prefetch_header(value: object) -> str:
    return _normalize_text(value).lower()


def _jinja_default(value: Any, default_value: Any = "", boolean: bool = False) -> Any:
    if value is None:
        return default_value
    if isinstance(value, str) and not value.strip():
        return default_value
    if boolean and not value:
        return default_value
    return value


class OutreachService:
    def __init__(
        self,
        *,
        db: Db,
        logger: logging.Logger,
        outreach_cfg: OutreachConfig,
        email_cfg: EmailConfig,
        officer_auth_cfg: OfficerAuthConfig,
        mailer: GraphMailer | None,
    ):
        self.db = db
        self.logger = logger
        self.cfg = outreach_cfg
        self.email_cfg = email_cfg
        self.officer_auth_cfg = officer_auth_cfg
        self.mailer = mailer
        self.env = Environment(
            undefined=ChainableUndefined,
            autoescape=False,
            finalize=lambda value: "" if value is None else value,
        )
        self.env.filters["default"] = _jinja_default
        self.html_env = Environment(
            undefined=ChainableUndefined,
            autoescape=True,
            finalize=lambda value: "" if value is None else value,
        )
        self.html_env.filters["default"] = _jinja_default

    def placeholder_catalog(self) -> list[str]:
        return list(_PLACEHOLDER_CATALOG)

    def send_enabled(self) -> bool:
        return bool(self.cfg.enabled and self.mailer and _normalize_text(self.cfg.sender_user_id))

    def send_readiness(self) -> dict[str, Any]:
        issues: list[str] = []
        public_base_url = _public_base_url(self.cfg, self.email_cfg, self.officer_auth_cfg)
        if not self.cfg.enabled:
            issues.append("outreach sending is disabled")
        if not _normalize_text(self.cfg.sender_user_id):
            issues.append("outreach sender mailbox is not configured")
        if self.mailer is None:
            issues.append("outreach Graph mailer is unavailable")
        if not public_base_url:
            issues.append("outreach public base URL is not configured")
        return {
            "enabled": bool(self.cfg.enabled),
            "ready": not issues,
            "sender_user_id": _normalize_text(self.cfg.sender_user_id) or None,
            "public_base_url": public_base_url or None,
            "reply_to_address": _normalize_text(self.cfg.reply_to_address) or None,
            "dry_run": bool(self.email_cfg.dry_run),
            "issues": issues,
        }

    async def ensure_seed_data(self) -> None:
        now = utcnow()
        async with aiosqlite.connect(self.db.db_path) as con:
            await con.execute("BEGIN IMMEDIATE")
            for template in _SEEDED_TEMPLATES:
                existing = await con.execute(
                    "SELECT id FROM outreach_templates WHERE template_key=?",
                    (template["template_key"],),
                )
                if await existing.fetchone():
                    continue
                await con.execute(
                    """
                    INSERT INTO outreach_templates(
                      template_key, name, template_type, subject_template, body_template,
                      active, seeded, created_at_utc, updated_at_utc
                    ) VALUES(?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        template["template_key"],
                        template["name"],
                        template["template_type"],
                        template["subject_template"],
                        template["body_template"],
                        1,
                        1,
                        now,
                        now,
                    ),
                )
            timezone_name = _normalize_text(self.cfg.timezone) or "America/New_York"
            for stop in _SEEDED_STOPS:
                existing = await con.execute(
                    "SELECT id FROM outreach_stops WHERE location_name=? AND visit_date_local=?",
                    (stop["location_name"], stop["visit_date_local"]),
                )
                if await existing.fetchone():
                    continue
                await con.execute(
                """
                INSERT INTO outreach_stops(
                  location_name, visit_date_local, start_time_local, end_time_local, timezone,
                      audience_location, audience_work_group, audience_group_name, audience_subgroup_name, notice_subject, reminder_subject,
                      notice_send_at_utc, reminder_send_at_utc, status, created_at_utc, updated_at_utc
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        stop["location_name"],
                        stop["visit_date_local"],
                        stop["start_time_local"],
                        stop["end_time_local"],
                        timezone_name,
                        stop["location_name"],
                        None,
                        None,
                        None,
                        stop["notice_subject"],
                        stop["reminder_subject"],
                        _default_notice_send_at(visit_date_local=stop["visit_date_local"], tz_name=timezone_name),
                        _default_reminder_send_at(
                            visit_date_local=stop["visit_date_local"],
                            start_time_local=stop["start_time_local"],
                            tz_name=timezone_name,
                        ),
                        "draft",
                        now,
                        now,
                    ),
                )
            await con.commit()

    async def list_contacts(self) -> list[dict[str, Any]]:
        rows = await self.db.fetchall(
            """
            SELECT id, email, first_name, last_name, full_name, work_location, work_group,
                   group_name, subgroup_name, department, bargaining_unit, local_number, steward_name, rep_name,
                   membership_type, employment_status, status_detail, status_bucket, status_source_text,
                   active, notes, source, extra_fields_json, created_at_utc, updated_at_utc
            FROM outreach_contacts
            ORDER BY lower(COALESCE(full_name, '')), lower(email)
            """
        )
        return [self._contact_row(row) for row in rows]

    async def save_contact(self, *, contact_id: int | None, payload: dict[str, Any]) -> dict[str, Any]:
        email = _normalize_email(payload.get("email"))
        if not email or "@" not in email:
            raise RuntimeError("valid email is required")
        existing_contact: dict[str, Any] | None = None
        if contact_id is not None:
            existing_contact = await self.get_contact(int(contact_id))
        else:
            existing_row = await self.db.fetchone("SELECT id FROM outreach_contacts WHERE email=?", (email,))
            if existing_row:
                existing_contact = await self.get_contact(int(existing_row[0]))

        def _merged_text(field_name: str) -> str:
            if field_name in payload:
                return _normalize_text(payload.get(field_name))
            if existing_contact:
                return _normalize_text(existing_contact.get(field_name))
            return ""

        first_name = _merged_text("first_name")
        last_name = _merged_text("last_name")
        full_name_input = _normalize_text(payload.get("full_name")) if "full_name" in payload else ""
        if full_name_input:
            full_name = full_name_input
        elif existing_contact and "full_name" not in payload and "first_name" not in payload and "last_name" not in payload:
            full_name = _normalize_text(existing_contact.get("full_name")) or _full_name(first_name, last_name, email)
        else:
            full_name = _full_name(first_name, last_name, email)

        extra_fields = payload.get("extra_fields") if "extra_fields" in payload else existing_contact.get("extra_fields") if existing_contact else {}
        if not isinstance(extra_fields, dict):
            extra_fields = {}
        cleaned_extra = {
            _normalize_key(key): _normalize_text(value)
            for key, value in extra_fields.items()
            if _normalize_key(key) and _normalize_text(value)
        }
        active = (
            _as_bool(payload.get("active"), default=True)
            if "active" in payload
            else bool(existing_contact["active"])
            if existing_contact
            else True
        )
        membership_type = _merged_text("membership_type") or None
        employment_status = _merged_text("employment_status") or None
        status_detail = _merged_text("status_detail") or None
        status_bucket = _merged_text("status_bucket") or None
        status_source_text = _merged_text("status_source_text") or None
        group_name = _merged_text("group_name") or _merged_text("work_group") or None
        subgroup_name = _merged_text("subgroup_name") or _merged_text("department") or None
        now = utcnow()
        if contact_id is None:
            await self.db.exec(
                """
                INSERT INTO outreach_contacts(
                  email, first_name, last_name, full_name, work_location, work_group, group_name, subgroup_name,
                  department, bargaining_unit, local_number, steward_name, rep_name, membership_type,
                  employment_status, status_detail, status_bucket, status_source_text, active, notes,
                  source, extra_fields_json, created_at_utc, updated_at_utc
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(email) DO UPDATE SET
                  first_name=excluded.first_name,
                  last_name=excluded.last_name,
                  full_name=excluded.full_name,
                  work_location=excluded.work_location,
                  work_group=excluded.work_group,
                  group_name=excluded.group_name,
                  subgroup_name=excluded.subgroup_name,
                  department=excluded.department,
                  bargaining_unit=excluded.bargaining_unit,
                  local_number=excluded.local_number,
                  steward_name=excluded.steward_name,
                  rep_name=excluded.rep_name,
                  membership_type=excluded.membership_type,
                  employment_status=excluded.employment_status,
                  status_detail=excluded.status_detail,
                  status_bucket=excluded.status_bucket,
                  status_source_text=excluded.status_source_text,
                  active=excluded.active,
                  notes=excluded.notes,
                  source=excluded.source,
                  extra_fields_json=excluded.extra_fields_json,
                  updated_at_utc=excluded.updated_at_utc
                """,
                (
                    email,
                    first_name,
                    last_name,
                    full_name,
                    _merged_text("work_location"),
                    _merged_text("work_group"),
                    group_name,
                    subgroup_name,
                    _merged_text("department"),
                    _merged_text("bargaining_unit"),
                    _merged_text("local_number"),
                    _merged_text("steward_name"),
                    _merged_text("rep_name"),
                    membership_type,
                    employment_status,
                    status_detail,
                    status_bucket,
                    status_source_text,
                    1 if active else 0,
                    _merged_text("notes"),
                    _merged_text("source") or "manual",
                    json.dumps(cleaned_extra, ensure_ascii=False),
                    now,
                    now,
                ),
            )
            row = await self.db.fetchone("SELECT id FROM outreach_contacts WHERE email=?", (email,))
            contact_id = int(row[0])
        else:
            await self.db.exec(
                """
                UPDATE outreach_contacts
                SET email=?, first_name=?, last_name=?, full_name=?, work_location=?, work_group=?, group_name=?, subgroup_name=?,
                    department=?, bargaining_unit=?, local_number=?, steward_name=?, rep_name=?,
                    membership_type=?, employment_status=?, status_detail=?, status_bucket=?, status_source_text=?,
                    active=?, notes=?, source=?, extra_fields_json=?, updated_at_utc=?
                WHERE id=?
                """,
                (
                    email,
                    first_name,
                    last_name,
                    full_name,
                    _merged_text("work_location"),
                    _merged_text("work_group"),
                    group_name,
                    subgroup_name,
                    _merged_text("department"),
                    _merged_text("bargaining_unit"),
                    _merged_text("local_number"),
                    _merged_text("steward_name"),
                    _merged_text("rep_name"),
                    membership_type,
                    employment_status,
                    status_detail,
                    status_bucket,
                    status_source_text,
                    1 if active else 0,
                    _merged_text("notes"),
                    _merged_text("source") or "manual",
                    json.dumps(cleaned_extra, ensure_ascii=False),
                    now,
                    int(contact_id),
                ),
            )
        return await self.get_contact(int(contact_id))

    async def get_contact(self, contact_id: int) -> dict[str, Any]:
        row = await self.db.fetchone(
            """
            SELECT id, email, first_name, last_name, full_name, work_location, work_group,
                   group_name, subgroup_name, department, bargaining_unit, local_number, steward_name, rep_name,
                   membership_type, employment_status, status_detail, status_bucket, status_source_text,
                   active, notes, source, extra_fields_json, created_at_utc, updated_at_utc
            FROM outreach_contacts
            WHERE id=?
            """,
            (int(contact_id),),
        )
        if not row:
            raise RuntimeError("contact not found")
        return self._contact_row(row)

    async def delete_contact(self, contact_id: int) -> None:
        await self.db.exec("DELETE FROM outreach_contacts WHERE id=?", (int(contact_id),))

    async def inspect_contacts_import(
        self,
        *,
        filename: str,
        content_base64: str,
        sheet_name: str | None = None,
        mapping: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        try:
            decoded = base64.b64decode(content_base64)
        except Exception as exc:
            raise RuntimeError("invalid import payload") from exc
        name = _normalize_text(filename).lower()
        sheets = self._import_sheets(name=name, content=decoded)
        selected_sheet = self._selected_import_sheet(sheets=sheets, requested_sheet_name=sheet_name)
        suggested_mapping = self._suggest_import_mapping(selected_sheet)
        remembered_mapping = await self._saved_import_mapping(headers=selected_sheet.headers)
        effective_mapping = self._resolve_import_mapping(
            headers=selected_sheet.headers,
            suggested_mapping=suggested_mapping,
            remembered_mapping=remembered_mapping,
            requested_mapping=mapping,
        )
        preview = await self._preview_import(sheet=selected_sheet, mapping=effective_mapping)
        return {
            "sheets": [
                {
                    "name": sheet.name,
                    "row_count": len(sheet.rows),
                    "selected": sheet.name == selected_sheet.name,
                }
                for sheet in sheets
            ],
            "selected_sheet_name": selected_sheet.name,
            "headers": list(selected_sheet.headers),
            "sample_rows": list(selected_sheet.rows[:5]),
            "suggested_mapping": suggested_mapping,
            "remembered_mapping": remembered_mapping,
            "effective_mapping": effective_mapping,
            "preview": preview,
        }

    async def import_contacts(
        self,
        *,
        filename: str,
        content_base64: str,
        sheet_name: str | None = None,
        mapping: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        inspect = await self.inspect_contacts_import(
            filename=filename,
            content_base64=content_base64,
            sheet_name=sheet_name,
            mapping=mapping,
        )
        decoded = base64.b64decode(content_base64)
        name = _normalize_text(filename).lower()
        sheets = self._import_sheets(name=name, content=decoded)
        selected_sheet = self._selected_import_sheet(
            sheets=sheets,
            requested_sheet_name=_normalize_text(inspect["selected_sheet_name"]) or sheet_name,
        )
        effective_mapping = dict(inspect["effective_mapping"])
        plan = await self._planned_import_rows(sheet=selected_sheet, mapping=effective_mapping)
        errors: list[str] = []
        for row in plan["rows"]:
            try:
                await self.save_contact(contact_id=row["existing_id"], payload=row["payload"])
            except Exception as exc:
                errors.append(f"Row {row['row_no']}: {exc}")
        if not errors:
            await self._save_import_mapping(headers=selected_sheet.headers, mapping=effective_mapping)
        response = dict(plan["preview"])
        response["selected_sheet_name"] = selected_sheet.name
        response["errors"] = errors
        response["saved_mapping"] = not errors
        return response

    def _import_sheets(self, *, name: str, content: bytes) -> list[OutreachImportSheet]:
        if name.endswith(".csv"):
            return [self._parse_csv_sheet(content)]
        if name.endswith(".xlsx"):
            return self._parse_xlsx_sheets(content)
        raise RuntimeError("only .csv and .xlsx imports are supported")

    def _parse_csv_sheet(self, content: bytes) -> OutreachImportSheet:
        text = content.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(StringIO(text)))
        if not rows:
            return OutreachImportSheet(name="CSV", headers=[], rows=[])
        headers = [str(value or "").strip() for value in rows[0]]
        parsed_rows: list[dict[str, str]] = []
        for values in rows[1:]:
            parsed_rows.append(
                {
                    header: _normalize_text(value)
                    for header, value in zip(headers, values, strict=False)
                    if header
                }
            )
        return OutreachImportSheet(name="CSV", headers=headers, rows=parsed_rows)

    def _parse_xlsx_sheets(self, content: bytes) -> list[OutreachImportSheet]:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        sheets: list[OutreachImportSheet] = []
        for worksheet in workbook.worksheets:
            values = list(worksheet.iter_rows(values_only=True))
            if not values:
                sheets.append(OutreachImportSheet(name=worksheet.title, headers=[], rows=[]))
                continue
            headers = [str(value or "").strip() for value in values[0]]
            parsed_rows: list[dict[str, str]] = []
            for row_values in values[1:]:
                parsed_rows.append(
                    {
                        header: _normalize_text(value)
                        for header, value in zip(headers, row_values, strict=False)
                        if header
                    }
                )
            sheets.append(OutreachImportSheet(name=worksheet.title, headers=headers, rows=parsed_rows))
        workbook.close()
        if not sheets:
            raise RuntimeError("the workbook does not contain any sheets")
        return sheets

    def _selected_import_sheet(
        self,
        *,
        sheets: list[OutreachImportSheet],
        requested_sheet_name: str | None,
    ) -> OutreachImportSheet:
        if not sheets:
            raise RuntimeError("the import file does not contain any rows")
        requested = _normalize_text(requested_sheet_name)
        if requested:
            for sheet in sheets:
                if sheet.name == requested:
                    return sheet
            raise RuntimeError(f"sheet not found: {requested}")
        return sheets[0]

    def _suggest_import_mapping(self, sheet: OutreachImportSheet) -> dict[str, Any]:
        field_mapping = {field_name: None for field_name in _IMPORT_FIELD_KEYS}
        for header in sheet.headers:
            normalized_header = _normalize_key(header)
            target = _CONTACT_HEADER_ALIASES.get(normalized_header)
            if target in field_mapping and field_mapping[target] is None:
                field_mapping[target] = header

        used_headers = {header for header in field_mapping.values() if header}

        def _sample_values(header: str | None) -> list[str]:
            if not header:
                return []
            return [_normalize_text(row.get(header)) for row in sheet.rows[:10]]

        def _find_header_by_alias(alias_set: set[str]) -> str | None:
            for header in sheet.headers:
                if header in used_headers:
                    continue
                if _normalize_key(header) in alias_set:
                    return header
            return None

        def _find_header_by_values(checker) -> str | None:  # noqa: ANN001
            for header in sheet.headers:
                if header in used_headers:
                    continue
                if any(checker(value) for value in _sample_values(header)):
                    return header
            return None

        def _find_structural_header(priority_headers: tuple[str, ...]) -> str | None:
            fallback: str | None = None
            for normalized_header in priority_headers:
                for header in sheet.headers:
                    if header in used_headers:
                        continue
                    if _normalize_key(header) != normalized_header:
                        continue
                    nonblank_count, unique_count = _header_value_stats(sheet.rows, header)
                    if nonblank_count <= 0:
                        continue
                    if unique_count > 1:
                        return header
                    if fallback is None:
                        fallback = header
            return fallback

        if field_mapping["group_name"] is None:
            field_mapping["group_name"] = _find_structural_header(_GROUP_HEADER_PRIORITY)
            if field_mapping["group_name"]:
                used_headers.add(field_mapping["group_name"])

        if field_mapping["subgroup_name"] is None:
            field_mapping["subgroup_name"] = _find_structural_header(_SUBGROUP_HEADER_PRIORITY)
            if field_mapping["subgroup_name"]:
                used_headers.add(field_mapping["subgroup_name"])

        if field_mapping["group_name"] is None and field_mapping["work_group"] is not None:
            field_mapping["group_name"] = field_mapping["work_group"]
        if field_mapping["group_name"] is None and field_mapping["department"] is not None:
            field_mapping["group_name"] = field_mapping["department"]
        if (
            field_mapping["subgroup_name"] is None
            and field_mapping["department"] is not None
            and field_mapping["department"] != field_mapping["group_name"]
        ):
            field_mapping["subgroup_name"] = field_mapping["department"]

        combined_header = _find_header_by_alias(_COMBINED_STATUS_HEADER_ALIASES)
        if combined_header is None:
            combined_header = _find_header_by_values(lambda value: bool(_classify_combined_status(value)["status_bucket"]))

        membership_header = _find_header_by_alias(_MEMBERSHIP_STATUS_HEADER_ALIASES)
        if membership_header:
            used_headers.add(membership_header)
        employment_header = _find_header_by_alias(_EMPLOYMENT_STATUS_HEADER_ALIASES)
        if employment_header:
            used_headers.add(employment_header)
        detail_header = _find_header_by_alias(_STATUS_DETAIL_HEADER_ALIASES)

        if membership_header is None:
            membership_header = _find_header_by_values(lambda value: bool(_canonical_membership_type(value)))
            if membership_header:
                used_headers.add(membership_header)
        if employment_header is None:
            employment_header = _find_header_by_values(lambda value: bool(_canonical_employment_status(value)))
            if employment_header:
                used_headers.add(employment_header)
        if detail_header is None:
            detail_header = _find_header_by_values(lambda value: bool(_canonical_status_detail(value)))

        mode = "combined" if combined_header else "split" if membership_header or employment_header or detail_header else "combined"
        return {
            "field_mapping": field_mapping,
            "status_mapping": {
                "mode": mode,
                "combined_status_column": combined_header,
                "membership_type_column": membership_header,
                "employment_status_column": employment_header,
                "status_detail_column": detail_header,
            },
        }

    async def _saved_import_mapping(self, *, headers: list[str]) -> dict[str, Any] | None:
        fingerprint = _header_fingerprint(headers)
        row = await self.db.fetchone(
            """
            SELECT mapping_json
            FROM outreach_import_profiles
            WHERE header_fingerprint=?
            LIMIT 1
            """,
            (fingerprint,),
        )
        if not row:
            return None
        mapping = _json_loads(row[0])
        if not mapping:
            return None
        return self._sanitize_import_mapping(headers=headers, mapping=mapping)

    def _sanitize_import_mapping(self, *, headers: list[str], mapping: dict[str, Any] | None) -> dict[str, Any]:
        valid_headers = set(headers)
        field_mapping_raw = _sanitize_mapping_dict((mapping or {}).get("field_mapping"))
        status_mapping_raw = _sanitize_mapping_dict((mapping or {}).get("status_mapping"))
        field_mapping = {
            field_name: header if header in valid_headers else None
            for field_name, header in {field_name: field_mapping_raw.get(field_name) for field_name in _IMPORT_FIELD_KEYS}.items()
        }
        mode = _normalize_key(status_mapping_raw.get("mode")) or "combined"
        if mode not in {"combined", "split"}:
            mode = "combined"
        status_mapping = {
            "mode": mode,
            "combined_status_column": status_mapping_raw.get("combined_status_column") if status_mapping_raw.get("combined_status_column") in valid_headers else None,
            "membership_type_column": status_mapping_raw.get("membership_type_column") if status_mapping_raw.get("membership_type_column") in valid_headers else None,
            "employment_status_column": status_mapping_raw.get("employment_status_column") if status_mapping_raw.get("employment_status_column") in valid_headers else None,
            "status_detail_column": status_mapping_raw.get("status_detail_column") if status_mapping_raw.get("status_detail_column") in valid_headers else None,
        }
        return {"field_mapping": field_mapping, "status_mapping": status_mapping}

    def _merge_import_mapping(self, base_mapping: dict[str, Any], override_mapping: dict[str, Any]) -> dict[str, Any]:
        merged = {
            "field_mapping": dict(base_mapping.get("field_mapping") or {}),
            "status_mapping": dict(base_mapping.get("status_mapping") or {}),
        }
        merged["field_mapping"].update({key: value for key, value in (override_mapping.get("field_mapping") or {}).items()})
        merged["status_mapping"].update({key: value for key, value in (override_mapping.get("status_mapping") or {}).items()})
        return merged

    def _resolve_import_mapping(
        self,
        *,
        headers: list[str],
        suggested_mapping: dict[str, Any],
        remembered_mapping: dict[str, Any] | None,
        requested_mapping: dict[str, Any] | None,
    ) -> dict[str, Any]:
        effective = self._sanitize_import_mapping(headers=headers, mapping=suggested_mapping)
        if remembered_mapping:
            effective = self._merge_import_mapping(effective, self._sanitize_import_mapping(headers=headers, mapping=remembered_mapping))
        if requested_mapping:
            effective = self._merge_import_mapping(effective, self._sanitize_import_mapping(headers=headers, mapping=requested_mapping))
        return effective

    async def _preview_import(self, *, sheet: OutreachImportSheet, mapping: dict[str, Any]) -> dict[str, Any]:
        plan = await self._planned_import_rows(sheet=sheet, mapping=mapping)
        return dict(plan["preview"])

    async def _planned_import_rows(self, *, sheet: OutreachImportSheet, mapping: dict[str, Any]) -> dict[str, Any]:
        existing_rows = await self.db.fetchall("SELECT id, email FROM outreach_contacts")
        existing_by_email = {
            _normalize_email(row[1]): int(row[0])
            for row in existing_rows
            if _normalize_email(row[1])
        }
        rows_to_commit: list[dict[str, Any]] = []
        preview = {
            "imported_count": 0,
            "updated_count": 0,
            "skipped_count": 0,
            "ignored_count": 0,
            "bucket_counts": {},
            "skipped_reasons": {},
            "ignored_reasons": {},
        }
        seen_emails: set[str] = set()
        for row_no, row in enumerate(sheet.rows, start=2):
            payload = self._mapped_contact_payload(row=row, mapping=mapping, source=sheet.name)
            email = _normalize_email(payload.get("email"))
            if not email:
                preview["skipped_count"] += 1
                preview["skipped_reasons"]["missing_email"] = preview["skipped_reasons"].get("missing_email", 0) + 1
                continue
            status_info = self._status_from_row(row=row, mapping=mapping)
            if status_info["reason"]:
                preview["ignored_count"] += 1
                reason = str(status_info["reason"])
                preview["ignored_reasons"][reason] = preview["ignored_reasons"].get(reason, 0) + 1
                continue
            if email in seen_emails:
                preview["skipped_count"] += 1
                preview["skipped_reasons"]["duplicate_email"] = preview["skipped_reasons"].get("duplicate_email", 0) + 1
                continue
            seen_emails.add(email)
            payload.update(
                {
                    "membership_type": status_info["membership_type"],
                    "employment_status": status_info["employment_status"],
                    "status_detail": status_info["status_detail"],
                    "status_bucket": status_info["status_bucket"],
                    "status_source_text": status_info["status_source_text"],
                }
            )
            existing_id = existing_by_email.get(email)
            if existing_id is not None:
                preview["updated_count"] += 1
            else:
                preview["imported_count"] += 1
            bucket = str(status_info["status_bucket"])
            preview["bucket_counts"][bucket] = preview["bucket_counts"].get(bucket, 0) + 1
            rows_to_commit.append({"row_no": row_no, "payload": payload, "existing_id": existing_id})
        return {"rows": rows_to_commit, "preview": preview}

    def _mapped_contact_payload(self, *, row: dict[str, str], mapping: dict[str, Any], source: str) -> dict[str, Any]:
        payload: dict[str, Any] = {"source": f"import:{_normalize_text(source) or 'sheet'}"}
        field_mapping = mapping.get("field_mapping") or {}
        for field_name in _IMPORT_FIELD_KEYS:
            header = _normalize_text(field_mapping.get(field_name))
            if not header:
                continue
            payload[field_name] = _normalize_text(row.get(header))
        return payload

    def _status_from_row(self, *, row: dict[str, str], mapping: dict[str, Any]) -> dict[str, str | None]:
        status_mapping = mapping.get("status_mapping") or {}
        mode = _normalize_key(status_mapping.get("mode")) or "combined"
        if mode == "split":
            membership_header = _normalize_text(status_mapping.get("membership_type_column"))
            employment_header = _normalize_text(status_mapping.get("employment_status_column"))
            detail_header = _normalize_text(status_mapping.get("status_detail_column"))
            if not (membership_header or employment_header or detail_header):
                return {
                    "membership_type": None,
                    "employment_status": None,
                    "status_detail": None,
                    "status_bucket": None,
                    "status_source_text": None,
                    "reason": "status_unmapped",
                }
            return _classify_split_status(
                membership_type_value=row.get(membership_header) if membership_header else "",
                employment_status_value=row.get(employment_header) if employment_header else "",
                status_detail_value=row.get(detail_header) if detail_header else "",
            )
        combined_header = _normalize_text(status_mapping.get("combined_status_column"))
        if not combined_header:
            return {
                "membership_type": None,
                "employment_status": None,
                "status_detail": None,
                "status_bucket": None,
                "status_source_text": None,
                "reason": "status_unmapped",
            }
        return _classify_combined_status(row.get(combined_header))

    async def _save_import_mapping(self, *, headers: list[str], mapping: dict[str, Any]) -> None:
        await self.db.exec(
            """
            INSERT INTO outreach_import_profiles(
              header_fingerprint, normalized_headers_json, mapping_json, created_at_utc, updated_at_utc
            ) VALUES(?,?,?,?,?)
            ON CONFLICT(header_fingerprint) DO UPDATE SET
              normalized_headers_json=excluded.normalized_headers_json,
              mapping_json=excluded.mapping_json,
              updated_at_utc=excluded.updated_at_utc
            """,
            (
                _header_fingerprint(headers),
                json.dumps(_normalized_headers(headers), ensure_ascii=False),
                json.dumps(mapping, ensure_ascii=False),
                utcnow(),
                utcnow(),
            ),
        )

    async def list_templates(self) -> list[dict[str, Any]]:
        rows = await self.db.fetchall(
            """
            SELECT id, template_key, name, template_type, subject_template, body_template,
                   active, seeded, created_at_utc, updated_at_utc
            FROM outreach_templates
            ORDER BY template_type, lower(name)
            """
        )
        return [self._template_row(row) for row in rows]

    async def save_template(self, *, template_id: int | None, payload: dict[str, Any]) -> dict[str, Any]:
        now = utcnow()
        if template_id is None:
            await self.db.exec(
                """
                INSERT INTO outreach_templates(
                  template_key, name, template_type, subject_template, body_template,
                  active, seeded, created_at_utc, updated_at_utc
                ) VALUES(?,?,?,?,?,?,?,?,?)
                ON CONFLICT(template_key) DO UPDATE SET
                  name=excluded.name,
                  template_type=excluded.template_type,
                  subject_template=excluded.subject_template,
                  body_template=excluded.body_template,
                  active=excluded.active,
                  updated_at_utc=excluded.updated_at_utc
                """,
                (
                    _normalize_key(payload.get("template_key")),
                    _normalize_text(payload.get("name")),
                    _normalize_key(payload.get("template_type")),
                    _normalize_text(payload.get("subject_template")),
                    _normalize_text(payload.get("body_template")),
                    1 if _as_bool(payload.get("active"), default=True) else 0,
                    0,
                    now,
                    now,
                ),
            )
            row = await self.db.fetchone(
                "SELECT id FROM outreach_templates WHERE template_key=?",
                (_normalize_key(payload.get("template_key")),),
            )
            template_id = int(row[0])
        else:
            await self.db.exec(
                """
                UPDATE outreach_templates
                SET template_key=?, name=?, template_type=?, subject_template=?, body_template=?,
                    active=?, updated_at_utc=?
                WHERE id=?
                """,
                (
                    _normalize_key(payload.get("template_key")),
                    _normalize_text(payload.get("name")),
                    _normalize_key(payload.get("template_type")),
                    _normalize_text(payload.get("subject_template")),
                    _normalize_text(payload.get("body_template")),
                    1 if _as_bool(payload.get("active"), default=True) else 0,
                    now,
                    int(template_id),
                ),
            )
        return await self.get_template(int(template_id))

    async def get_template(self, template_id: int) -> dict[str, Any]:
        row = await self.db.fetchone(
            """
            SELECT id, template_key, name, template_type, subject_template, body_template,
                   active, seeded, created_at_utc, updated_at_utc
            FROM outreach_templates
            WHERE id=?
            """,
            (int(template_id),),
        )
        if not row:
            raise RuntimeError("template not found")
        return self._template_row(row)

    async def delete_template(self, template_id: int) -> None:
        await self.db.exec("DELETE FROM outreach_templates WHERE id=?", (int(template_id),))

    async def list_stops(self) -> list[dict[str, Any]]:
        rows = await self.db.fetchall(
            """
            SELECT id, location_name, visit_date_local, start_time_local, end_time_local,
                   timezone, audience_location, audience_work_group, audience_group_name, audience_subgroup_name,
                   audience_status_bucket, notice_subject, reminder_subject, notice_send_at_utc, reminder_send_at_utc,
                   status, created_at_utc, updated_at_utc
            FROM outreach_stops
            ORDER BY visit_date_local, start_time_local, lower(location_name)
            """
        )
        return [self._stop_row(row) for row in rows]

    async def save_stop(self, *, stop_id: int | None, payload: dict[str, Any]) -> dict[str, Any]:
        timezone_name = _normalize_text(payload.get("timezone")) or _normalize_text(self.cfg.timezone) or "America/New_York"
        visit_date_local = _normalize_text(payload.get("visit_date_local"))
        start_time_local = _normalize_text(payload.get("start_time_local"))
        end_time_local = _normalize_text(payload.get("end_time_local"))
        notice_send_at_local = _normalize_text(payload.get("notice_send_at_local"))
        reminder_send_at_local = _normalize_text(payload.get("reminder_send_at_local"))
        notice_send_at_utc = (
            _local_to_utc_iso(
                tz_name=timezone_name,
                local_date=notice_send_at_local.split("T", 1)[0],
                local_time_value=_parse_local_time(notice_send_at_local.split("T", 1)[1]),
            )
            if notice_send_at_local
            else _default_notice_send_at(visit_date_local=visit_date_local, tz_name=timezone_name)
        )
        reminder_send_at_utc = (
            _local_to_utc_iso(
                tz_name=timezone_name,
                local_date=reminder_send_at_local.split("T", 1)[0],
                local_time_value=_parse_local_time(reminder_send_at_local.split("T", 1)[1]),
            )
            if reminder_send_at_local
            else _default_reminder_send_at(
                visit_date_local=visit_date_local,
                start_time_local=start_time_local,
                tz_name=timezone_name,
            )
        )
        audience_work_group = _normalize_text_list_value(payload.get("audience_work_group"))
        audience_group_name = _normalize_text_list_value(payload.get("audience_group_name")) or audience_work_group
        audience_subgroup_name = _normalize_text_list_value(payload.get("audience_subgroup_name"))
        now = utcnow()
        if stop_id is None:
            await self.db.exec(
                """
                INSERT INTO outreach_stops(
                  location_name, visit_date_local, start_time_local, end_time_local, timezone,
                  audience_location, audience_work_group, audience_group_name, audience_subgroup_name, audience_status_bucket, notice_subject,
                  reminder_subject, notice_send_at_utc, reminder_send_at_utc, status,
                  created_at_utc, updated_at_utc
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(location_name, visit_date_local) DO UPDATE SET
                  start_time_local=excluded.start_time_local,
                  end_time_local=excluded.end_time_local,
                  timezone=excluded.timezone,
                  audience_location=excluded.audience_location,
                  audience_work_group=excluded.audience_work_group,
                  audience_group_name=excluded.audience_group_name,
                  audience_subgroup_name=excluded.audience_subgroup_name,
                  audience_status_bucket=excluded.audience_status_bucket,
                  notice_subject=excluded.notice_subject,
                  reminder_subject=excluded.reminder_subject,
                  notice_send_at_utc=excluded.notice_send_at_utc,
                  reminder_send_at_utc=excluded.reminder_send_at_utc,
                  status=excluded.status,
                  updated_at_utc=excluded.updated_at_utc
                """,
                (
                    _normalize_text(payload.get("location_name")),
                    visit_date_local,
                    start_time_local,
                    end_time_local,
                    timezone_name,
                    _normalize_text(payload.get("audience_location")),
                    audience_work_group,
                    audience_group_name,
                    audience_subgroup_name or None,
                    _normalize_text(payload.get("audience_status_bucket")) or None,
                    _normalize_text(payload.get("notice_subject")),
                    _normalize_text(payload.get("reminder_subject")),
                    notice_send_at_utc,
                    reminder_send_at_utc,
                    _normalize_key(payload.get("status")) or "draft",
                    now,
                    now,
                ),
            )
            row = await self.db.fetchone(
                "SELECT id FROM outreach_stops WHERE location_name=? AND visit_date_local=?",
                (_normalize_text(payload.get("location_name")), visit_date_local),
            )
            stop_id = int(row[0])
        else:
            await self.db.exec(
                """
                UPDATE outreach_stops
                SET location_name=?, visit_date_local=?, start_time_local=?, end_time_local=?, timezone=?,
                    audience_location=?, audience_work_group=?, audience_group_name=?, audience_subgroup_name=?,
                    audience_status_bucket=?, notice_subject=?,
                    reminder_subject=?, notice_send_at_utc=?, reminder_send_at_utc=?, status=?, updated_at_utc=?
                WHERE id=?
                """,
                (
                    _normalize_text(payload.get("location_name")),
                    visit_date_local,
                    start_time_local,
                    end_time_local,
                    timezone_name,
                    _normalize_text(payload.get("audience_location")),
                    audience_work_group,
                    audience_group_name,
                    audience_subgroup_name or None,
                    _normalize_text(payload.get("audience_status_bucket")) or None,
                    _normalize_text(payload.get("notice_subject")),
                    _normalize_text(payload.get("reminder_subject")),
                    notice_send_at_utc,
                    reminder_send_at_utc,
                    _normalize_key(payload.get("status")) or "draft",
                    now,
                    int(stop_id),
                ),
            )
        return await self.get_stop(int(stop_id))

    async def get_stop(self, stop_id: int) -> dict[str, Any]:
        row = await self.db.fetchone(
            """
            SELECT id, location_name, visit_date_local, start_time_local, end_time_local,
                   timezone, audience_location, audience_work_group, audience_group_name, audience_subgroup_name,
                   audience_status_bucket, notice_subject, reminder_subject, notice_send_at_utc, reminder_send_at_utc,
                   status, created_at_utc, updated_at_utc
            FROM outreach_stops
            WHERE id=?
            """,
            (int(stop_id),),
        )
        if not row:
            raise RuntimeError("stop not found")
        return self._stop_row(row)

    async def delete_stop(self, stop_id: int) -> None:
        await self.db.exec("DELETE FROM outreach_stops WHERE id=?", (int(stop_id),))

    async def list_suppressions(self) -> list[dict[str, Any]]:
        rows = await self.db.fetchall(
            """
            SELECT id, email, contact_id, reason, created_at_utc
            FROM outreach_suppressions
            ORDER BY created_at_utc DESC, id DESC
            """
        )
        return [
            {
                "id": int(row[0]),
                "email": _normalize_text(row[1]),
                "contact_id": int(row[2]) if row[2] is not None else None,
                "reason": _normalize_text(row[3]),
                "created_at_utc": _normalize_text(row[4]),
            }
            for row in rows
        ]

    async def delete_suppression(self, suppression_id: int) -> None:
        await self.db.exec("DELETE FROM outreach_suppressions WHERE id=?", (int(suppression_id),))

    async def list_send_log(self, *, limit: int = 200) -> list[dict[str, Any]]:
        rows = await self.db.fetchall(
            """
            SELECT l.id, l.stop_id, l.template_id, l.contact_id, l.recipient_email, l.email_type,
                   l.subject, l.status, l.scheduled_for_utc, l.attempted_at_utc, l.sent_at_utc,
                   l.failed_at_utc, l.graph_message_id, l.internet_message_id, l.error_text,
                   s.location_name, s.visit_date_local, l.created_at_utc
            FROM outreach_send_log l
            LEFT JOIN outreach_stops s ON s.id=l.stop_id
            ORDER BY l.created_at_utc DESC, l.id DESC
            LIMIT ?
            """,
            (int(limit),),
        )
        return [
            {
                "id": int(row[0]),
                "stop_id": int(row[1]) if row[1] is not None else None,
                "template_id": int(row[2]) if row[2] is not None else None,
                "contact_id": int(row[3]) if row[3] is not None else None,
                "recipient_email": _normalize_text(row[4]),
                "email_type": _normalize_text(row[5]),
                "subject": _normalize_text(row[6]),
                "status": _normalize_text(row[7]),
                "scheduled_for_utc": _normalize_text(row[8]) or None,
                "attempted_at_utc": _normalize_text(row[9]) or None,
                "sent_at_utc": _normalize_text(row[10]) or None,
                "failed_at_utc": _normalize_text(row[11]) or None,
                "graph_message_id": _normalize_text(row[12]) or None,
                "internet_message_id": _normalize_text(row[13]) or None,
                "error_text": _normalize_text(row[14]) or None,
                "location_name": _normalize_text(row[15]) or None,
                "visit_date_local": _normalize_text(row[16]) or None,
                "created_at_utc": _normalize_text(row[17]),
            }
            for row in rows
        ]

    async def summary(self) -> dict[str, int]:
        sent_row = await self.db.fetchone("SELECT COUNT(*) FROM outreach_send_log WHERE status='sent'")
        failed_row = await self.db.fetchone("SELECT COUNT(*) FROM outreach_send_log WHERE status='failed'")
        suppression_row = await self.db.fetchone("SELECT COUNT(*) FROM outreach_suppressions")
        stop_row = await self.db.fetchone("SELECT COUNT(*) FROM outreach_stops")
        active_contact_row = await self.db.fetchone("SELECT COUNT(*) FROM outreach_contacts WHERE active=1")
        return {
            "sent_count": int(sent_row[0] if sent_row else 0),
            "failed_count": int(failed_row[0] if failed_row else 0),
            "suppressed_count": int(suppression_row[0] if suppression_row else 0),
            "stop_count": int(stop_row[0] if stop_row else 0),
            "active_contact_count": int(active_contact_row[0] if active_contact_row else 0),
        }

    def tracking_pixel_bytes(self) -> bytes:
        return _PIXEL_GIF_BYTES

    async def record_click(
        self,
        token: str,
        *,
        client_ip: str | None = None,
        user_agent: str | None = None,
        purpose: str | None = None,
    ) -> str | None:
        hashed = _token_hash(_normalize_text(token))
        row = await self.db.fetchone(
            """
            SELECT id, send_log_id, destination_url
            FROM outreach_tracked_links
            WHERE tracking_token_hash=?
            ORDER BY id DESC
            LIMIT 1
            """,
            (hashed,),
        )
        if not row:
            return None
        tracked_link_id = int(row[0])
        send_log_id = int(row[1])
        destination_url = _normalize_text(row[2])
        await self._log_outreach_event(
            send_log_id=send_log_id,
            tracked_link_id=tracked_link_id,
            event_type="click",
            client_ip=client_ip,
            user_agent=user_agent,
            purpose=purpose,
        )
        return destination_url

    async def record_open(
        self,
        token: str,
        *,
        client_ip: str | None = None,
        user_agent: str | None = None,
        purpose: str | None = None,
    ) -> bool:
        hashed = _token_hash(_normalize_text(token))
        row = await self.db.fetchone(
            """
            SELECT id
            FROM outreach_send_log
            WHERE open_token_hash=?
            ORDER BY id DESC
            LIMIT 1
            """,
            (hashed,),
        )
        if not row:
            return False
        await self._log_outreach_event(
            send_log_id=int(row[0]),
            tracked_link_id=None,
            event_type="estimated_open",
            client_ip=client_ip,
            user_agent=user_agent,
            purpose=purpose,
        )
        return True

    async def analytics_dashboard(
        self,
        *,
        date_from: str | None = None,
        date_to: str | None = None,
        stop_id: int | None = None,
        location: str | None = None,
        template_id: int | None = None,
        recipient_email: str | None = None,
        work_group: str | None = None,
        limit: int = 50,
    ) -> dict[str, Any]:
        filtered = await self._analytics_rows(
            date_from=date_from,
            date_to=date_to,
            stop_id=stop_id,
            location=location,
            template_id=template_id,
            recipient_email=recipient_email,
            work_group=work_group,
        )
        send_rows = filtered["send_rows"]
        event_rows = filtered["event_rows"]
        top_links = self._top_clicked_links(event_rows)
        by_stop = self._campaign_summary(send_rows, event_rows)
        totals = self._analytics_totals(send_rows, event_rows)
        return {
            "filters": {
                "date_from": date_from,
                "date_to": date_to,
                "stop_id": stop_id,
                "location": location,
                "template_id": template_id,
                "recipient_email": recipient_email,
                "work_group": work_group,
            },
            "totals": totals,
            "top_links": top_links[:10],
            "campaigns": by_stop,
            "recent_activity": event_rows[: max(1, int(limit))],
            "notes": [
                "Estimated opens are image-load estimates, not guaranteed human reads.",
                "Clicks are generally more reliable than opens.",
                "Metrics exclude only clearly flagged automation or prefetch traffic when identified.",
            ],
        }

    async def export_campaign_summary_csv(self, **filters: Any) -> str:
        dashboard = await self.analytics_dashboard(**filters)
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(
            [
                "stop_id",
                "location",
                "visit_date",
                "template_count",
                "sent",
                "failed",
                "suppressed",
                "unsubscribes",
                "estimated_opens",
                "unique_estimated_opens",
                "clicks",
                "unique_clicks",
            ]
        )
        for row in dashboard["campaigns"]:
            writer.writerow(
                [
                    row.get("stop_id") or "",
                    row.get("location_name") or "",
                    row.get("visit_date_local") or "",
                    row.get("template_count") or 0,
                    row.get("sent_count") or 0,
                    row.get("failed_count") or 0,
                    row.get("suppressed_count") or 0,
                    row.get("unsubscribe_count") or 0,
                    row.get("estimated_open_count") or 0,
                    row.get("unique_estimated_open_count") or 0,
                    row.get("click_count") or 0,
                    row.get("unique_click_count") or 0,
                ]
            )
        return out.getvalue()

    async def export_click_activity_csv(self, **filters: Any) -> str:
        filtered = await self._analytics_rows(**filters)
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(
            [
                "occurred_at_utc",
                "recipient_email",
                "location_name",
                "visit_date_local",
                "destination_url",
                "suspected_automation",
                "automation_reason",
            ]
        )
        for row in filtered["event_rows"]:
            if row["event_type"] != "click":
                continue
            writer.writerow(
                [
                    row["occurred_at_utc"],
                    row.get("recipient_email") or "",
                    row.get("location_name") or "",
                    row.get("visit_date_local") or "",
                    row.get("destination_url") or "",
                    "yes" if row.get("suspected_automation") else "no",
                    row.get("automation_reason") or "",
                ]
            )
        return out.getvalue()

    async def export_send_history_csv(self, **filters: Any) -> str:
        filtered = await self._analytics_rows(**filters)
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(
            [
                "send_log_id",
                "recipient_email",
                "email_type",
                "subject",
                "status",
                "template_name",
                "location_name",
                "visit_date_local",
                "work_group",
                "sent_at_utc",
                "failed_at_utc",
            ]
        )
        for row in filtered["send_rows"]:
            writer.writerow(
                [
                    row["id"],
                    row["recipient_email"],
                    row["email_type"],
                    row["subject"],
                    row["status"],
                    row.get("template_name") or "",
                    row.get("location_name") or "",
                    row.get("visit_date_local") or "",
                    row.get("work_group") or "",
                    row.get("sent_at_utc") or "",
                    row.get("failed_at_utc") or "",
                ]
            )
        return out.getvalue()

    async def export_suppressions_csv(self) -> str:
        rows = await self.list_suppressions()
        out = StringIO()
        writer = csv.writer(out)
        writer.writerow(["email", "contact_id", "reason", "created_at_utc"])
        for row in rows:
            writer.writerow([row["email"], row.get("contact_id") or "", row["reason"], row["created_at_utc"]])
        return out.getvalue()

    async def preview(self, *, template_id: int, stop_id: int, contact_id: int | None, recipient_email: str | None) -> OutreachRenderedMessage:
        template_row = await self._resolve_template_row(template_id)
        stop_row = await self._resolve_stop_row(stop_id)
        contact_row = await self.get_contact(contact_id) if contact_id is not None else None
        return self._render_message(
            template_row=template_row,
            stop_row=stop_row,
            contact_row=contact_row,
            recipient_email=recipient_email or (contact_row["email"] if contact_row else ""),
            unsubscribe_url=self._preview_unsubscribe_url(),
        )

    async def preview_one_off(
        self,
        *,
        template_id: int,
        stop_id: int,
        recipient_email: str,
        contact_id: int | None = None,
        manual_contact: dict[str, Any] | None = None,
    ) -> OutreachRenderedMessage:
        template_row = await self._resolve_template_row(template_id)
        stop_row = await self._resolve_stop_row(stop_id)
        contact_row = await self._resolve_contact_row(
            contact_id=contact_id,
            recipient_email=recipient_email,
            manual_contact=manual_contact,
        )
        return self._render_message(
            template_row=template_row,
            stop_row=stop_row,
            contact_row=contact_row,
            recipient_email=recipient_email,
            unsubscribe_url=self._preview_unsubscribe_url(),
        )

    async def preview_quick_message(
        self,
        *,
        stop_id: int,
        recipient_email: str,
        subject_template: str,
        body_template: str,
        contact_id: int | None = None,
        manual_contact: dict[str, Any] | None = None,
    ) -> OutreachRenderedMessage:
        stop_row = await self._resolve_stop_row(stop_id)
        contact_row = await self._resolve_contact_row(
            contact_id=contact_id,
            recipient_email=recipient_email,
            manual_contact=manual_contact,
        )
        template_row = {
            "id": None,
            "template_key": "quick_test_message",
            "template_type": "ad_hoc",
            "subject_template": subject_template,
            "body_template": body_template,
        }
        return self._render_message(
            template_row=template_row,
            stop_row=stop_row,
            contact_row=contact_row,
            recipient_email=recipient_email,
            unsubscribe_url=self._preview_unsubscribe_url(),
        )

    async def send_test(
        self,
        *,
        template_id: int,
        stop_id: int,
        contact_id: int | None,
        recipient_email: str,
    ) -> OutreachSendSummary:
        template_row = await self._resolve_template_row(template_id)
        stop_row = await self._resolve_stop_row(stop_id)
        contact_row = await self.get_contact(contact_id) if contact_id is not None else None
        return await self._send(
            template_row=template_row,
            stop_row=stop_row,
            contact_row=contact_row,
            recipient_email=recipient_email,
            email_type="test",
            scheduled_for_utc=None,
        )

    async def send_test_quick_message(
        self,
        *,
        stop_id: int,
        recipient_email: str,
        subject_template: str,
        body_template: str,
        contact_id: int | None = None,
        manual_contact: dict[str, Any] | None = None,
    ) -> OutreachSendSummary:
        stop_row = await self._resolve_stop_row(stop_id)
        contact_row = await self._resolve_contact_row(
            contact_id=contact_id,
            recipient_email=recipient_email,
            manual_contact=manual_contact,
        )
        template_row = {
            "id": None,
            "template_key": "quick_test_message",
            "template_type": "ad_hoc",
            "subject_template": subject_template,
            "body_template": body_template,
        }
        return await self._send(
            template_row=template_row,
            stop_row=stop_row,
            contact_row=contact_row,
            recipient_email=recipient_email,
            email_type="test",
            scheduled_for_utc=None,
        )

    async def send_one_off(
        self,
        *,
        template_id: int,
        stop_id: int,
        recipient_email: str,
        contact_id: int | None = None,
        manual_contact: dict[str, Any] | None = None,
    ) -> OutreachSendSummary:
        template_row = await self._resolve_template_row(template_id)
        stop_row = await self._resolve_stop_row(stop_id)
        contact_row = await self._resolve_contact_row(
            contact_id=contact_id,
            recipient_email=recipient_email,
            manual_contact=manual_contact,
        )
        return await self._send(
            template_row=template_row,
            stop_row=stop_row,
            contact_row=contact_row,
            recipient_email=recipient_email,
            email_type="one_off",
            scheduled_for_utc=None,
        )

    async def send_test_one_off(
        self,
        *,
        template_id: int,
        stop_id: int,
        recipient_email: str,
        contact_id: int | None = None,
        manual_contact: dict[str, Any] | None = None,
    ) -> OutreachSendSummary:
        template_row = await self._resolve_template_row(template_id)
        stop_row = await self._resolve_stop_row(stop_id)
        contact_row = await self._resolve_contact_row(
            contact_id=contact_id,
            recipient_email=recipient_email,
            manual_contact=manual_contact,
        )
        return await self._send(
            template_row=template_row,
            stop_row=stop_row,
            contact_row=contact_row,
            recipient_email=recipient_email,
            email_type="test",
            scheduled_for_utc=None,
        )

    async def run_due(self) -> dict[str, Any]:
        now_utc = datetime.now(timezone.utc).isoformat()
        due_rows = await self.db.fetchall(
            """
            SELECT id, location_name, visit_date_local, start_time_local, end_time_local, timezone,
                   audience_location, audience_work_group, audience_group_name, audience_subgroup_name,
                   audience_status_bucket, notice_subject, reminder_subject, notice_send_at_utc,
                   reminder_send_at_utc, status, created_at_utc, updated_at_utc
            FROM outreach_stops
            WHERE status='active'
              AND (notice_send_at_utc<=? OR reminder_send_at_utc<=?)
            ORDER BY visit_date_local, start_time_local, id
            """,
            (now_utc, now_utc),
        )
        processed_count = 0
        sent_count = 0
        failed_count = 0
        skipped_suppressed_count = 0
        skipped_existing_count = 0
        results: list[OutreachSendSummary] = []
        for row in due_rows:
            stop_row = self._stop_row(row)
            due_types: list[tuple[str, str, str]] = []
            if _normalize_text(stop_row["notice_send_at_utc"]) <= now_utc:
                due_types.append(("notice", "notice_send_at_utc", "notice"))
            if _normalize_text(stop_row["reminder_send_at_utc"]) <= now_utc:
                due_types.append(("reminder", "reminder_send_at_utc", "reminder"))
            for email_type, scheduled_key, template_type in due_types:
                template_row = await self._template_for_type(template_type)
                contacts = await self._contacts_for_stop(stop_row)
                for contact_row in contacts:
                    if processed_count >= self.cfg.max_sends_per_run:
                        return {
                            "processed_count": processed_count,
                            "sent_count": sent_count,
                            "failed_count": failed_count,
                            "skipped_suppressed_count": skipped_suppressed_count,
                            "skipped_existing_count": skipped_existing_count,
                            "rows": [self._send_summary_row(item) for item in results],
                        }
                    processed_count += 1
                    if await self._is_suppressed(contact_row["email"]):
                        await self._log_suppressed_attempt(
                            template_row=template_row,
                            stop_row=stop_row,
                            contact_row=contact_row,
                            recipient_email=contact_row["email"],
                            email_type=email_type,
                            scheduled_for_utc=stop_row[scheduled_key],
                        )
                        skipped_suppressed_count += 1
                        continue
                    existing = await self.db.fetchone(
                        """
                        SELECT id FROM outreach_send_log
                        WHERE stop_id=? AND email_type=? AND recipient_email=?
                        """,
                        (int(stop_row["id"]), email_type, contact_row["email"]),
                    )
                    if existing:
                        skipped_existing_count += 1
                        continue
                    summary = await self._send(
                        template_row=template_row,
                        stop_row=stop_row,
                        contact_row=contact_row,
                        recipient_email=contact_row["email"],
                        email_type=email_type,
                        scheduled_for_utc=stop_row[scheduled_key],
                    )
                    results.append(summary)
                    if summary.status == "sent":
                        sent_count += 1
                    else:
                        failed_count += 1
                    if self.cfg.min_seconds_between_sends > 0:
                        await asyncio.sleep(self.cfg.min_seconds_between_sends)
        return {
            "processed_count": processed_count,
            "sent_count": sent_count,
            "failed_count": failed_count,
            "skipped_suppressed_count": skipped_suppressed_count,
            "skipped_existing_count": skipped_existing_count,
            "rows": [self._send_summary_row(item) for item in results],
        }

    async def unsubscribe(self, token: str) -> dict[str, str]:
        hashed = _token_hash(_normalize_text(token))
        row = await self.db.fetchone(
            """
            SELECT id, recipient_email, contact_id
            FROM outreach_send_log
            WHERE unsubscribe_token_hash=?
            ORDER BY id DESC
            LIMIT 1
            """,
            (hashed,),
        )
        if not row:
            raise RuntimeError("unsubscribe link is invalid or expired")
        send_log_id = int(row[0])
        email = _normalize_email(row[1])
        contact_id = int(row[2]) if row[2] is not None else None
        existing = await self.db.fetchone("SELECT id FROM outreach_suppressions WHERE email=?", (email,))
        if not existing:
            await self.db.exec(
                """
                INSERT INTO outreach_suppressions(email, contact_id, reason, created_at_utc)
                VALUES(?,?,?,?)
                """,
                (email, contact_id, "unsubscribe", utcnow()),
            )
        prior_event = await self.db.fetchone(
            "SELECT id FROM outreach_events WHERE send_log_id=? AND event_type='unsubscribe' LIMIT 1",
            (send_log_id,),
        )
        if not prior_event:
            await self._log_outreach_event(
                send_log_id=send_log_id,
                tracked_link_id=None,
                event_type="unsubscribe",
                client_ip=None,
                user_agent=None,
                purpose=None,
            )
        return {"email": email, "status": "suppressed", "reason": "unsubscribe"}

    async def _template_for_type(self, template_type: str) -> dict[str, Any]:
        row = await self.db.fetchone(
            """
            SELECT id, template_key, name, template_type, subject_template, body_template,
                   active, seeded, created_at_utc, updated_at_utc
            FROM outreach_templates
            WHERE template_type=? AND active=1
            ORDER BY seeded DESC, id ASC
            LIMIT 1
            """,
            (_normalize_key(template_type),),
        )
        if not row:
            raise RuntimeError(f"active outreach template not found for type {template_type}")
        return self._template_row(row)

    async def _contacts_for_stop(self, stop_row: dict[str, Any]) -> list[dict[str, Any]]:
        params: list[Any] = []
        sql = """
            SELECT id, email, first_name, last_name, full_name, work_location, work_group,
                   group_name, subgroup_name, department, bargaining_unit, local_number, steward_name, rep_name,
                   membership_type, employment_status, status_detail, status_bucket, status_source_text,
                   active, notes, source, extra_fields_json, created_at_utc, updated_at_utc
            FROM outreach_contacts
            WHERE active=1
        """
        if _normalize_text(stop_row["audience_location"]):
            sql += " AND lower(COALESCE(work_location, ''))=lower(?)"
            params.append(stop_row["audience_location"])
        work_group_filters = [token.lower() for token in _normalize_text_list(stop_row["audience_work_group"])]
        if work_group_filters:
            placeholders = ",".join("?" for _ in work_group_filters)
            sql += f" AND lower(COALESCE(work_group, '')) IN ({placeholders})"
            params.extend(work_group_filters)
        group_filters = [token.lower() for token in _normalize_text_list(stop_row.get("audience_group_name"))]
        if group_filters:
            placeholders = ",".join("?" for _ in group_filters)
            sql += f" AND lower(COALESCE(group_name, work_group, '')) IN ({placeholders})"
            params.extend(group_filters)
        subgroup_filters = [token.lower() for token in _normalize_text_list(stop_row.get("audience_subgroup_name"))]
        if subgroup_filters:
            placeholders = ",".join("?" for _ in subgroup_filters)
            sql += f" AND lower(COALESCE(subgroup_name, department, '')) IN ({placeholders})"
            params.extend(subgroup_filters)
        if _normalize_text(stop_row.get("audience_status_bucket")):
            sql += " AND lower(COALESCE(status_bucket, ''))=lower(?)"
            params.append(stop_row["audience_status_bucket"])
        sql += " ORDER BY lower(COALESCE(full_name, '')), lower(email)"
        rows = await self.db.fetchall(sql, tuple(params))
        return [self._contact_row(row) for row in rows]

    async def _is_suppressed(self, email: str) -> bool:
        row = await self.db.fetchone("SELECT id FROM outreach_suppressions WHERE email=?", (_normalize_email(email),))
        return bool(row)

    async def _log_suppressed_attempt(
        self,
        *,
        template_row: dict[str, Any],
        stop_row: dict[str, Any],
        contact_row: dict[str, Any],
        recipient_email: str,
        email_type: str,
        scheduled_for_utc: str | None,
    ) -> None:
        existing = await self.db.fetchone(
            """
            SELECT id FROM outreach_send_log
            WHERE stop_id=? AND email_type=? AND recipient_email=?
            """,
            (int(stop_row["id"]), email_type, _normalize_email(recipient_email)),
        )
        if existing:
            return
        rendered = self._render_message(
            template_row=template_row,
            stop_row=stop_row,
            contact_row=contact_row,
            recipient_email=recipient_email,
            unsubscribe_url=self._preview_unsubscribe_url(),
        )
        now = utcnow()
        await self.db.exec(
            """
            INSERT INTO outreach_send_log(
              stop_id, template_id, contact_id, recipient_email, email_type, subject,
              text_body, html_body, merge_data_json, scheduled_for_utc, attempted_at_utc,
              status, created_at_utc, updated_at_utc
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                int(stop_row["id"]),
                int(template_row["id"]),
                int(contact_row["id"]),
                _normalize_email(recipient_email),
                email_type,
                rendered.subject,
                rendered.text_body,
                rendered.html_body,
                json.dumps(
                    self._build_context(
                        contact_row=contact_row,
                        stop_row=stop_row,
                        recipient_email=recipient_email,
                        subject=rendered.subject,
                        unsubscribe_url=self._preview_unsubscribe_url(),
                    ),
                    ensure_ascii=False,
                ),
                scheduled_for_utc,
                now,
                "suppressed",
                now,
                now,
            ),
        )

    async def _create_tracked_links(
        self,
        *,
        send_log_id: int,
        text_body: str,
        html_body: str,
        unsubscribe_url: str,
        public_base_url: str,
    ) -> dict[str, str]:
        destinations = self._extract_trackable_urls(
            text_body=text_body,
            html_body=html_body,
            unsubscribe_url=unsubscribe_url,
            public_base_url=public_base_url,
        )
        replacements: dict[str, str] = {}
        for destination in destinations:
            raw_token = secrets.token_urlsafe(24)
            redirect_url = f"{public_base_url}/r/{raw_token}"
            await self.db.exec(
                """
                INSERT INTO outreach_tracked_links(send_log_id, destination_url, tracking_token_hash, link_label, created_at_utc)
                VALUES(?,?,?,?,?)
                """,
                (
                    int(send_log_id),
                    destination,
                    _token_hash(raw_token),
                    destination,
                    utcnow(),
                ),
            )
            replacements[destination] = redirect_url
        return replacements

    def _extract_trackable_urls(
        self,
        *,
        text_body: str,
        html_body: str,
        unsubscribe_url: str,
        public_base_url: str,
    ) -> list[str]:
        candidates: list[str] = []
        for match in _URL_RE.findall(text_body or ""):
            candidates.append(match)
        for match in _HTML_HREF_RE.findall(html_body or ""):
            candidates.append(match)
        out: list[str] = []
        seen: set[str] = set()
        for candidate in candidates:
            cleaned = _normalize_text(candidate).rstrip(".,);")
            parsed = urlsplit(cleaned)
            if parsed.scheme not in {"http", "https"}:
                continue
            if cleaned == unsubscribe_url or "/unsubscribe/" in cleaned:
                continue
            if cleaned.startswith(f"{public_base_url}/r/") or cleaned.startswith(f"{public_base_url}/o/"):
                continue
            if cleaned in seen:
                continue
            seen.add(cleaned)
            out.append(cleaned)
        return out

    def _rewrite_text_tracking_links(self, text_body: str, replacements: dict[str, str]) -> str:
        if not replacements:
            return text_body
        rewritten = text_body
        for destination, redirect_url in replacements.items():
            rewritten = rewritten.replace(destination, redirect_url)
        return rewritten

    def _rewrite_html_tracking_links(self, html_body: str, replacements: dict[str, str]) -> str:
        if not replacements:
            return html_body

        def _replace_href(match: re.Match[str]) -> str:
            original = _normalize_text(match.group(1))
            return f'href="{html.escape(replacements.get(original, original), quote=True)}"'

        return _HTML_HREF_RE.sub(_replace_href, html_body)

    def _append_open_tracking_pixel(self, html_body: str, pixel_url: str) -> str:
        pixel_html = (
            f'<img src="{html.escape(pixel_url, quote=True)}" alt="" width="1" height="1" '
            'style="display:block;border:0;width:1px;height:1px;" referrerpolicy="no-referrer" />'
        )
        if "</body>" in html_body:
            return html_body.replace("</body>", f"{pixel_html}</body>")
        return html_body + pixel_html

    async def _log_outreach_event(
        self,
        *,
        send_log_id: int,
        tracked_link_id: int | None,
        event_type: str,
        client_ip: str | None,
        user_agent: str | None,
        purpose: str | None,
    ) -> None:
        metadata = self._event_metadata(event_type=event_type, user_agent=user_agent, purpose=purpose)
        await self.db.exec(
            """
            INSERT INTO outreach_events(
              send_log_id, tracked_link_id, event_type, occurred_at_utc, ip_hash, user_agent_hash, metadata_json
            ) VALUES(?,?,?,?,?,?,?)
            """,
            (
                int(send_log_id),
                int(tracked_link_id) if tracked_link_id is not None else None,
                event_type,
                utcnow(),
                _hash_text(client_ip),
                _hash_text(_normalize_user_agent(user_agent)),
                json.dumps(metadata, ensure_ascii=False),
            ),
        )

    def _event_metadata(self, *, event_type: str, user_agent: str | None, purpose: str | None) -> dict[str, Any]:
        normalized_ua = _normalize_user_agent(user_agent).lower()
        normalized_purpose = _normalize_prefetch_header(purpose)
        automation_reason = ""
        if normalized_purpose in {"prefetch", "preview"}:
            automation_reason = f"request-purpose:{normalized_purpose}"
        else:
            for hint in _STRONG_AUTOMATION_HINTS:
                if hint in normalized_ua:
                    automation_reason = f"user-agent:{hint}"
                    break
        suspected_automation = bool(automation_reason)
        ignore_for_metrics = suspected_automation
        if event_type == "estimated_open" and automation_reason.startswith("user-agent:googleimageproxy"):
            ignore_for_metrics = False
        return {
            "event_type": event_type,
            "suspected_automation": suspected_automation,
            "automation_reason": automation_reason or None,
            "ignore_for_metrics": ignore_for_metrics,
        }

    async def _analytics_rows(
        self,
        *,
        date_from: str | None = None,
        date_to: str | None = None,
        stop_id: int | None = None,
        location: str | None = None,
        template_id: int | None = None,
        recipient_email: str | None = None,
        work_group: str | None = None,
    ) -> dict[str, list[dict[str, Any]]]:
        send_rows = await self.db.fetchall(
            """
            SELECT l.id, l.stop_id, l.template_id, l.contact_id, l.recipient_email, l.email_type,
                   l.subject, l.status, l.scheduled_for_utc, l.attempted_at_utc, l.sent_at_utc,
                   l.failed_at_utc, l.graph_message_id, l.internet_message_id, l.error_text,
                   l.created_at_utc, l.merge_data_json, s.location_name, s.visit_date_local,
                   t.name, t.template_key
            FROM outreach_send_log l
            LEFT JOIN outreach_stops s ON s.id=l.stop_id
            LEFT JOIN outreach_templates t ON t.id=l.template_id
            ORDER BY COALESCE(l.sent_at_utc, l.failed_at_utc, l.attempted_at_utc, l.created_at_utc) DESC, l.id DESC
            """
        )
        parsed_send_rows: list[dict[str, Any]] = []
        for row in send_rows:
            merge_data = _json_loads(row[16])
            parsed_send_rows.append(
                {
                    "id": int(row[0]),
                    "stop_id": int(row[1]) if row[1] is not None else None,
                    "template_id": int(row[2]) if row[2] is not None else None,
                    "contact_id": int(row[3]) if row[3] is not None else None,
                    "recipient_email": _normalize_email(row[4]),
                    "email_type": _normalize_text(row[5]),
                    "subject": _normalize_text(row[6]),
                    "status": _normalize_text(row[7]),
                    "scheduled_for_utc": _normalize_text(row[8]) or None,
                    "attempted_at_utc": _normalize_text(row[9]) or None,
                    "sent_at_utc": _normalize_text(row[10]) or None,
                    "failed_at_utc": _normalize_text(row[11]) or None,
                    "graph_message_id": _normalize_text(row[12]) or None,
                    "internet_message_id": _normalize_text(row[13]) or None,
                    "error_text": _normalize_text(row[14]) or None,
                    "created_at_utc": _normalize_text(row[15]),
                    "merge_data": merge_data,
                    "location_name": _normalize_text(row[17]) or merge_data.get("campaign_location") or None,
                    "visit_date_local": _normalize_text(row[18]) or None,
                    "template_name": _normalize_text(row[19]) or None,
                    "template_key": _normalize_text(row[20]) or None,
                    "work_group": _normalize_text(merge_data.get("work_group")) or None,
                }
            )
        filtered_send_rows = [
            row
            for row in parsed_send_rows
            if self._row_matches_analytics_filters(
                row,
                date_from=date_from,
                date_to=date_to,
                stop_id=stop_id,
                location=location,
                template_id=template_id,
                recipient_email=recipient_email,
                work_group=work_group,
            )
        ]
        allowed_send_ids = {row["id"] for row in filtered_send_rows}
        event_rows = await self.db.fetchall(
            """
            SELECT e.id, e.send_log_id, e.tracked_link_id, e.event_type, e.occurred_at_utc,
                   e.ip_hash, e.user_agent_hash, e.metadata_json, l.destination_url
            FROM outreach_events e
            LEFT JOIN outreach_tracked_links l ON l.id=e.tracked_link_id
            ORDER BY e.occurred_at_utc DESC, e.id DESC
            """
        )
        send_lookup = {row["id"]: row for row in filtered_send_rows}
        parsed_event_rows: list[dict[str, Any]] = []
        for row in event_rows:
            send_log_id = int(row[1])
            if send_log_id not in allowed_send_ids:
                continue
            send_row = send_lookup[send_log_id]
            metadata = _json_loads(row[7])
            occurred_at_utc = _normalize_text(row[4])
            if date_from and occurred_at_utc[:10] < date_from:
                continue
            if date_to and occurred_at_utc[:10] > date_to:
                continue
            parsed_event_rows.append(
                {
                    "id": int(row[0]),
                    "send_log_id": send_log_id,
                    "tracked_link_id": int(row[2]) if row[2] is not None else None,
                    "event_type": _normalize_text(row[3]),
                    "occurred_at_utc": occurred_at_utc,
                    "ip_hash": _normalize_text(row[5]) or None,
                    "user_agent_hash": _normalize_text(row[6]) or None,
                    "destination_url": _normalize_text(row[8]) or None,
                    "recipient_email": send_row["recipient_email"],
                    "location_name": send_row.get("location_name"),
                    "visit_date_local": send_row.get("visit_date_local"),
                    "template_name": send_row.get("template_name"),
                    "template_id": send_row.get("template_id"),
                    "work_group": send_row.get("work_group"),
                    "stop_id": send_row.get("stop_id"),
                    "suspected_automation": bool(metadata.get("suspected_automation")),
                    "automation_reason": _normalize_text(metadata.get("automation_reason")) or None,
                    "ignore_for_metrics": bool(metadata.get("ignore_for_metrics")),
                }
            )
        return {"send_rows": filtered_send_rows, "event_rows": parsed_event_rows}

    def _row_matches_analytics_filters(
        self,
        row: dict[str, Any],
        *,
        date_from: str | None,
        date_to: str | None,
        stop_id: int | None,
        location: str | None,
        template_id: int | None,
        recipient_email: str | None,
        work_group: str | None,
    ) -> bool:
        timestamp = (
            row.get("sent_at_utc")
            or row.get("failed_at_utc")
            or row.get("attempted_at_utc")
            or row.get("created_at_utc")
            or ""
        )
        if date_from and str(timestamp)[:10] < date_from:
            return False
        if date_to and str(timestamp)[:10] > date_to:
            return False
        if stop_id is not None and row.get("stop_id") != int(stop_id):
            return False
        if template_id is not None and row.get("template_id") != int(template_id):
            return False
        if location and _normalize_text(row.get("location_name")).lower() != _normalize_text(location).lower():
            return False
        if recipient_email and row.get("recipient_email") != _normalize_email(recipient_email):
            return False
        if work_group and _normalize_text(row.get("work_group")).lower() != _normalize_text(work_group).lower():
            return False
        return True

    def _analytics_totals(self, send_rows: list[dict[str, Any]], event_rows: list[dict[str, Any]]) -> dict[str, int]:
        metric_events = [row for row in event_rows if not row.get("ignore_for_metrics")]
        unique_open_send_ids = {row["send_log_id"] for row in metric_events if row["event_type"] == "estimated_open"}
        unique_click_pairs = {
            (row["send_log_id"], row.get("tracked_link_id"))
            for row in metric_events
            if row["event_type"] == "click"
        }
        unsubscribe_send_ids = {row["send_log_id"] for row in event_rows if row["event_type"] == "unsubscribe"}
        return {
            "sent_count": sum(1 for row in send_rows if row["status"] == "sent"),
            "failed_count": sum(1 for row in send_rows if row["status"] == "failed"),
            "suppressed_count": sum(1 for row in send_rows if row["status"] == "suppressed"),
            "unsubscribe_count": len(unsubscribe_send_ids),
            "estimated_open_count": sum(1 for row in metric_events if row["event_type"] == "estimated_open"),
            "unique_estimated_open_count": len(unique_open_send_ids),
            "click_count": sum(1 for row in metric_events if row["event_type"] == "click"),
            "unique_click_count": len(unique_click_pairs),
        }

    def _campaign_summary(self, send_rows: list[dict[str, Any]], event_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        campaign_map: dict[tuple[Any, ...], dict[str, Any]] = {}
        for row in send_rows:
            key = (row.get("stop_id"), row.get("location_name"), row.get("visit_date_local"))
            bucket = campaign_map.setdefault(
                key,
                {
                    "stop_id": row.get("stop_id"),
                    "location_name": row.get("location_name"),
                    "visit_date_local": row.get("visit_date_local"),
                    "template_count": 0,
                    "sent_count": 0,
                    "failed_count": 0,
                    "suppressed_count": 0,
                    "unsubscribe_count": 0,
                    "estimated_open_count": 0,
                    "unique_estimated_open_count": 0,
                    "click_count": 0,
                    "unique_click_count": 0,
                    "_template_ids": set(),
                    "_unique_open_send_ids": set(),
                    "_unique_click_pairs": set(),
                    "_unsubscribe_send_ids": set(),
                    "_send_ids": set(),
                },
            )
            bucket["_send_ids"].add(row["id"])
            if row.get("template_id") is not None:
                bucket["_template_ids"].add(row["template_id"])
            if row["status"] == "sent":
                bucket["sent_count"] += 1
            elif row["status"] == "failed":
                bucket["failed_count"] += 1
            elif row["status"] == "suppressed":
                bucket["suppressed_count"] += 1
        bucket_by_send_id: dict[int, dict[str, Any]] = {}
        for bucket in campaign_map.values():
            for send_id in bucket["_send_ids"]:
                bucket_by_send_id[int(send_id)] = bucket
        for event in event_rows:
            bucket = bucket_by_send_id.get(int(event["send_log_id"]))
            if bucket is None:
                continue
            if event["event_type"] == "unsubscribe":
                bucket["_unsubscribe_send_ids"].add(event["send_log_id"])
                continue
            if event.get("ignore_for_metrics"):
                continue
            if event["event_type"] == "estimated_open":
                bucket["estimated_open_count"] += 1
                bucket["_unique_open_send_ids"].add(event["send_log_id"])
            elif event["event_type"] == "click":
                bucket["click_count"] += 1
                bucket["_unique_click_pairs"].add((event["send_log_id"], event.get("tracked_link_id")))
        out: list[dict[str, Any]] = []
        for bucket in campaign_map.values():
            bucket["template_count"] = len(bucket["_template_ids"])
            bucket["unique_estimated_open_count"] = len(bucket["_unique_open_send_ids"])
            bucket["unique_click_count"] = len(bucket["_unique_click_pairs"])
            bucket["unsubscribe_count"] = len(bucket["_unsubscribe_send_ids"])
            for internal_key in ("_template_ids", "_unique_open_send_ids", "_unique_click_pairs", "_unsubscribe_send_ids", "_send_ids"):
                bucket.pop(internal_key, None)
            out.append(bucket)
        out.sort(
            key=lambda row: (
                _normalize_text(row.get("visit_date_local")),
                _normalize_text(row.get("location_name")),
            )
        )
        return out

    def _top_clicked_links(self, event_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        link_map: dict[str, dict[str, Any]] = {}
        for row in event_rows:
            if row["event_type"] != "click" or row.get("ignore_for_metrics"):
                continue
            destination = _normalize_text(row.get("destination_url"))
            if not destination:
                continue
            bucket = link_map.setdefault(
                destination,
                {
                    "destination_url": destination,
                    "click_count": 0,
                    "unique_click_count": 0,
                    "_unique_pairs": set(),
                },
            )
            bucket["click_count"] += 1
            bucket["_unique_pairs"].add((row["send_log_id"], row.get("tracked_link_id")))
        out: list[dict[str, Any]] = []
        for bucket in link_map.values():
            bucket["unique_click_count"] = len(bucket["_unique_pairs"])
            bucket.pop("_unique_pairs", None)
            out.append(bucket)
        out.sort(key=lambda row: (-int(row["click_count"]), row["destination_url"]))
        return out

    async def _resolve_contact_row(
        self,
        *,
        contact_id: int | None,
        recipient_email: str,
        manual_contact: dict[str, Any] | None,
    ) -> dict[str, Any] | None:
        base: dict[str, Any] | None = await self.get_contact(contact_id) if contact_id is not None else None
        manual = manual_contact or {}
        if not base and not any(_normalize_text(value) for key, value in manual.items() if key != "extra_fields"):
            extra = manual.get("extra_fields")
            if not isinstance(extra, dict) or not extra:
                return None
        merged = dict(base or {})
        merged["email"] = _normalize_email(recipient_email)
        for key in (
            "first_name",
            "last_name",
            "full_name",
            "work_location",
            "work_group",
            "group_name",
            "subgroup_name",
            "department",
            "bargaining_unit",
            "local_number",
            "steward_name",
            "rep_name",
            "membership_type",
            "employment_status",
            "status_detail",
            "status_bucket",
            "status_source_text",
            "source",
            "notes",
        ):
            value = manual.get(key)
            if value is not None and _normalize_text(value):
                merged[key] = value
        extra_fields = {}
        if isinstance(base, dict) and isinstance(base.get("extra_fields"), dict):
            extra_fields.update(base["extra_fields"])
        if isinstance(manual.get("extra_fields"), dict):
            for key, value in manual["extra_fields"].items():
                if _normalize_text(key):
                    extra_fields[str(key)] = str(value or "")
        if extra_fields:
            merged["extra_fields"] = extra_fields
        return merged

    async def _resolve_stop_row(self, stop_id: int | None) -> dict[str, Any]:
        if stop_id is not None and int(stop_id) > 0:
            return await self.get_stop(int(stop_id))
        rows = await self.list_stops()
        if not rows:
            raise RuntimeError("no outreach stops are available")
        return rows[0]

    async def _resolve_template_row(self, template_id: int | None) -> dict[str, Any]:
        if template_id is not None and int(template_id) > 0:
            return await self.get_template(int(template_id))
        rows = await self.list_templates()
        if not rows:
            raise RuntimeError("no outreach templates are available")
        return rows[0]

    async def _send(
        self,
        *,
        template_row: dict[str, Any],
        stop_row: dict[str, Any],
        contact_row: dict[str, Any] | None,
        recipient_email: str,
        email_type: str,
        scheduled_for_utc: str | None,
    ) -> OutreachSendSummary:
        if not self.send_enabled():
            raise RuntimeError("outreach sending is not enabled in config")
        raw_token = secrets.token_urlsafe(32)
        public_base_url = _public_base_url(self.cfg, self.email_cfg, self.officer_auth_cfg)
        if not public_base_url:
            raise RuntimeError("outreach.public_base_url or an equivalent public app URL must be configured")
        unsubscribe_url = f"{public_base_url}/unsubscribe/{raw_token}"
        open_token = secrets.token_urlsafe(24)
        open_pixel_url = f"{public_base_url}/o/{open_token}.gif"
        rendered = self._render_message(
            template_row=template_row,
            stop_row=stop_row,
            contact_row=contact_row,
            recipient_email=recipient_email,
            unsubscribe_url=unsubscribe_url,
        )
        now = utcnow()
        template_id = template_row.get("id")
        template_key = _normalize_text(template_row.get("template_key")) or "outreach_manual"
        send_log_id = await self.db.insert(
            """
            INSERT INTO outreach_send_log(
              stop_id, template_id, contact_id, recipient_email, email_type, subject,
              text_body, html_body, merge_data_json, scheduled_for_utc, attempted_at_utc,
              status, unsubscribe_token_hash, open_token_hash, created_at_utc, updated_at_utc
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                int(stop_row["id"]),
                int(template_id) if template_id is not None else None,
                int(contact_row["id"]) if contact_row and contact_row.get("id") is not None else None,
                _normalize_email(recipient_email),
                email_type,
                rendered.subject,
                rendered.text_body,
                rendered.html_body,
                json.dumps(
                    self._build_context(
                        contact_row=contact_row,
                        stop_row=stop_row,
                        recipient_email=recipient_email,
                        subject=rendered.subject,
                        unsubscribe_url=unsubscribe_url,
                    ),
                    ensure_ascii=False,
                ),
                scheduled_for_utc,
                now,
                "pending",
                _token_hash(raw_token),
                _token_hash(open_token),
                now,
                now,
            ),
        )
        tracked_links = await self._create_tracked_links(
            send_log_id=int(send_log_id),
            text_body=rendered.text_body,
            html_body=rendered.html_body,
            unsubscribe_url=unsubscribe_url,
            public_base_url=public_base_url,
        )
        tracked_text_body = self._rewrite_text_tracking_links(rendered.text_body, tracked_links)
        tracked_html_body = self._rewrite_html_tracking_links(rendered.html_body, tracked_links)
        tracked_html_body = self._append_open_tracking_pixel(tracked_html_body, open_pixel_url)
        await self.db.exec(
            """
            UPDATE outreach_send_log
            SET text_body=?, html_body=?, updated_at_utc=?
            WHERE id=?
            """,
            (tracked_text_body, tracked_html_body, utcnow(), int(send_log_id)),
        )
        mailto_target = _normalize_text(self.cfg.reply_to_address or self.email_cfg.derek_email)
        list_unsubscribe = [f"<{unsubscribe_url}>"]
        if mailto_target:
            list_unsubscribe.append(f"<mailto:{quote(mailto_target, safe='@')}?subject=unsubscribe>")
        try:
            sent = self.mailer.send_mime_mail(
                to_recipients=[recipient_email],
                subject=rendered.subject,
                text_body=tracked_text_body,
                html_body=tracked_html_body,
                custom_headers={
                    "List-Unsubscribe": ", ".join(list_unsubscribe),
                    "List-Unsubscribe-Post": "List-Unsubscribe=One-Click",
                    "X-Outreach-Stop-ID": str(stop_row["id"]),
                    "X-Outreach-Email-Type": email_type,
                    "X-Outreach-Template-Key": template_key,
                },
                from_display_name=_normalize_text(self.cfg.sender_display_name) or None,
                reply_to_address=_normalize_text(self.cfg.reply_to_address) or None,
                reply_to_name=_normalize_text(self.cfg.reply_to_name) or None,
            )
        except Exception as exc:
            error_text = str(exc)
            await self.db.exec(
                """
                UPDATE outreach_send_log
                SET status='failed', failed_at_utc=?, error_text=?, updated_at_utc=?
                WHERE id=?
                """,
                (utcnow(), error_text, utcnow(), int(send_log_id)),
            )
            return OutreachSendSummary(
                send_log_id=int(send_log_id),
                recipient_email=_normalize_email(recipient_email),
                status="failed",
                graph_message_id=None,
                error_text=error_text,
            )
        await self.db.exec(
            """
            UPDATE outreach_send_log
            SET status='sent', sent_at_utc=?, graph_message_id=?, internet_message_id=?, updated_at_utc=?
            WHERE id=?
            """,
            (utcnow(), sent.graph_message_id, sent.internet_message_id, utcnow(), int(send_log_id)),
        )
        return OutreachSendSummary(
            send_log_id=int(send_log_id),
            recipient_email=_normalize_email(recipient_email),
            status="sent",
            graph_message_id=sent.graph_message_id,
            error_text=None,
        )

    def _preview_unsubscribe_url(self) -> str:
        public_base_url = _public_base_url(self.cfg, self.email_cfg, self.officer_auth_cfg)
        if public_base_url:
            return f"{public_base_url}/unsubscribe/preview"
        return "https://example.invalid/unsubscribe/preview"

    def _render_message(
        self,
        *,
        template_row: dict[str, Any],
        stop_row: dict[str, Any],
        contact_row: dict[str, Any] | None,
        recipient_email: str,
        unsubscribe_url: str,
    ) -> OutreachRenderedMessage:
        context = self._build_context(
            contact_row=contact_row,
            stop_row=stop_row,
            recipient_email=recipient_email,
            subject="",
            unsubscribe_url=unsubscribe_url,
        )
        subject_template = self._subject_template(template_row=template_row, stop_row=stop_row)
        unknown_fields = self._unknown_placeholders(subject_template, context) | self._unknown_placeholders(
            template_row["body_template"],
            context,
        )
        subject = self.env.from_string(subject_template).render(context).strip()
        context["subject"] = subject
        body = self.env.from_string(template_row["body_template"]).render(context).strip()
        text_body = self._text_with_footer(body, unsubscribe_url)
        html_body = self._html_from_text(body, unsubscribe_url, context=context)
        return OutreachRenderedMessage(
            subject=subject,
            text_body=text_body,
            html_body=html_body,
            unknown_placeholders=sorted(unknown_fields),
        )

    def _subject_template(self, *, template_row: dict[str, Any], stop_row: dict[str, Any]) -> str:
        if template_row["template_type"] == "notice" and _normalize_text(stop_row["notice_subject"]):
            return stop_row["notice_subject"]
        if template_row["template_type"] == "reminder" and _normalize_text(stop_row["reminder_subject"]):
            return stop_row["reminder_subject"]
        return template_row["subject_template"]

    def _unknown_placeholders(self, template_source: str, context: dict[str, Any]) -> set[str]:
        ast = self.env.parse(template_source or "")
        return {
            variable
            for variable in meta.find_undeclared_variables(ast)
            if variable not in context
        }

    def _build_context(
        self,
        *,
        contact_row: dict[str, Any] | None,
        stop_row: dict[str, Any],
        recipient_email: str,
        subject: str,
        unsubscribe_url: str,
    ) -> dict[str, Any]:
        contact = contact_row or {}
        first_name = _normalize_text(contact.get("first_name"))
        last_name = _normalize_text(contact.get("last_name"))
        full_name = _normalize_text(contact.get("full_name")) or _full_name(first_name, last_name, recipient_email)
        visit_date = _friendly_visit_date(stop_row["visit_date_local"])
        visit_time = _friendly_visit_time(stop_row["start_time_local"], stop_row["end_time_local"])
        context: dict[str, Any] = {
            "first_name": first_name,
            "last_name": last_name,
            "full_name": full_name,
            "email": _normalize_email(recipient_email),
            "location": stop_row["location_name"],
            "campaign_location": stop_row["location_name"],
            "work_location": _normalize_text(contact.get("work_location")),
            "work_group": _normalize_text(contact.get("work_group")),
            "group_name": _normalize_text(contact.get("group_name")) or _normalize_text(contact.get("work_group")),
            "subgroup_name": _normalize_text(contact.get("subgroup_name")) or _normalize_text(contact.get("department")),
            "department": _normalize_text(contact.get("department")),
            "bargaining_unit": _normalize_text(contact.get("bargaining_unit")),
            "local_number": _normalize_text(contact.get("local_number")),
            "steward_name": _normalize_text(contact.get("steward_name")),
            "rep_name": _normalize_text(contact.get("rep_name")),
            "membership_type": _normalize_text(contact.get("membership_type")),
            "employment_status": _normalize_text(contact.get("employment_status")),
            "status_detail": _normalize_text(contact.get("status_detail")),
            "status_bucket": _normalize_text(contact.get("status_bucket")),
            "status_source_text": _normalize_text(contact.get("status_source_text")),
            "visit_date": visit_date,
            "visit_time": visit_time,
            "subject": subject,
            "sender_name": _normalize_text(self.cfg.sender_display_name),
            "reply_to": _normalize_text(self.cfg.reply_to_address),
            "unsubscribe_url": unsubscribe_url,
        }
        extra_fields = contact.get("extra_fields")
        if isinstance(extra_fields, dict):
            for key, value in extra_fields.items():
                normalized_key = _normalize_key(key)
                if normalized_key and normalized_key not in context:
                    context[normalized_key] = value
        return context

    def _text_with_footer(self, body: str, unsubscribe_url: str) -> str:
        return f"{body.rstrip()}\n\nTo unsubscribe, visit:\n{unsubscribe_url}\n"

    def _html_from_text(
        self,
        body: str,
        unsubscribe_url: str,
        *,
        context: dict[str, Any] | None = None,
    ) -> str:
        content_html = Markup(self._body_content_html(body))
        wrapper_context = dict(context or {})
        wrapper_context.update(
            {
                "body_text": body.rstrip(),
                "body_html": content_html,
                "content_html": content_html,
                "unsubscribe_url": unsubscribe_url,
            }
        )
        custom_template_path = Path(self.email_cfg.templates_dir) / _OUTREACH_HTML_TEMPLATE_FILENAME
        if custom_template_path.exists():
            try:
                template_source = custom_template_path.read_text(encoding="utf-8")
                return self.html_env.from_string(template_source).render(wrapper_context).strip()
            except Exception:
                self.logger.exception("outreach html wrapper render failed", extra={"template_path": str(custom_template_path)})
        return self.html_env.from_string(_DEFAULT_OUTREACH_HTML_TEMPLATE).render(wrapper_context).strip()

    def _body_content_html(self, body: str) -> str:
        paragraphs = []
        for block in re.split(r"\n{2,}", body.strip()):
            block = block.strip()
            if not block:
                continue
            escaped = self._linkify_plain_text_html(block).replace("\n", "<br>")
            paragraphs.append(f"<p>{escaped}</p>")
        return "".join(paragraphs)

    def _linkify_plain_text_html(self, text: str) -> str:
        rendered: list[str] = []
        cursor = 0
        for match in _URL_RE.finditer(text):
            rendered.append(html.escape(text[cursor:match.start()]))
            url = match.group(1)
            safe_url = html.escape(url, quote=True)
            safe_label = html.escape(url)
            rendered.append(f'<a href="{safe_url}" style="color:#0f766e;text-decoration:underline;">{safe_label}</a>')
            cursor = match.end()
        rendered.append(html.escape(text[cursor:]))
        return "".join(rendered)

    def _contact_row(self, row: tuple[Any, ...]) -> dict[str, Any]:
        return {
            "id": int(row[0]),
            "email": _normalize_email(row[1]),
            "first_name": _normalize_text(row[2]) or None,
            "last_name": _normalize_text(row[3]) or None,
            "full_name": _normalize_text(row[4]),
            "work_location": _normalize_text(row[5]) or None,
            "work_group": _normalize_text(row[6]) or None,
            "group_name": _normalize_text(row[7]) or _normalize_text(row[6]) or None,
            "subgroup_name": _normalize_text(row[8]) or _normalize_text(row[9]) or None,
            "department": _normalize_text(row[9]) or None,
            "bargaining_unit": _normalize_text(row[10]) or None,
            "local_number": _normalize_text(row[11]) or None,
            "steward_name": _normalize_text(row[12]) or None,
            "rep_name": _normalize_text(row[13]) or None,
            "membership_type": _normalize_text(row[14]) or None,
            "employment_status": _normalize_text(row[15]) or None,
            "status_detail": _normalize_text(row[16]) or None,
            "status_bucket": _normalize_text(row[17]) or None,
            "status_source_text": _normalize_text(row[18]) or None,
            "active": bool(int(row[19] or 0)),
            "notes": _normalize_text(row[20]) or None,
            "source": _normalize_text(row[21]) or "manual",
            "extra_fields": _json_loads(row[22]),
            "created_at_utc": _normalize_text(row[23]),
            "updated_at_utc": _normalize_text(row[24]),
        }

    def _template_row(self, row: tuple[Any, ...]) -> dict[str, Any]:
        return {
            "id": int(row[0]),
            "template_key": _normalize_text(row[1]),
            "name": _normalize_text(row[2]),
            "template_type": _normalize_text(row[3]),
            "subject_template": _normalize_text(row[4]),
            "body_template": _normalize_text(row[5]),
            "active": bool(int(row[6] or 0)),
            "seeded": bool(int(row[7] or 0)),
            "created_at_utc": _normalize_text(row[8]),
            "updated_at_utc": _normalize_text(row[9]),
        }

    def _stop_row(self, row: tuple[Any, ...]) -> dict[str, Any]:
        timezone_name = _normalize_text(row[5]) or "America/New_York"
        return {
            "id": int(row[0]),
            "location_name": _normalize_text(row[1]),
            "visit_date_local": _normalize_text(row[2]),
            "start_time_local": _normalize_text(row[3]),
            "end_time_local": _normalize_text(row[4]),
            "timezone": timezone_name,
            "audience_location": _normalize_text(row[6]) or None,
            "audience_work_group": _normalize_text(row[7]) or None,
            "audience_group_name": _normalize_text(row[8]) or _normalize_text(row[7]) or None,
            "audience_subgroup_name": _normalize_text(row[9]) or None,
            "audience_status_bucket": _normalize_text(row[10]) or None,
            "notice_subject": _normalize_text(row[11]) or None,
            "reminder_subject": _normalize_text(row[12]) or None,
            "notice_send_at_utc": _normalize_text(row[13]),
            "reminder_send_at_utc": _normalize_text(row[14]),
            "status": _normalize_text(row[15]) or "draft",
            "created_at_utc": _normalize_text(row[16]),
            "updated_at_utc": _normalize_text(row[17]),
            "notice_send_at_local": _utc_to_local_input(_normalize_text(row[13]), timezone_name),
            "reminder_send_at_local": _utc_to_local_input(_normalize_text(row[14]), timezone_name),
        }

    def _send_summary_row(self, row: OutreachSendSummary) -> dict[str, Any]:
        return {
            "send_log_id": row.send_log_id,
            "recipient_email": row.recipient_email,
            "status": row.status,
            "graph_message_id": row.graph_message_id,
            "error_text": row.error_text,
        }
